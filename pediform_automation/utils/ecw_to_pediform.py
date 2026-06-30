"""
utils/ecw_to_pediform.py
------------------------
Converts the ECW Encounter Patient Download into a PediForm-compatible
import Excel file.

Key rules:
  - Dates MUST be stored as plain text strings (not Excel date serials).
    PediForm rejects date serial numbers with "Patient DOB is required".
  - Only columns PediForm needs are written; all ECW extras are dropped.
  - Only eligible patients (filtered by excel_reader) are included.
  - Appointment column = "MM/DD/YYYY HH:MM" as text.
"""

from datetime import datetime, date, time as dt_time
from typing import List, Dict, Optional

import openpyxl
from openpyxl.styles import Font, numbers as xl_numbers

from config.settings import PEDIFORM_IMPORT_PATH
from utils.logger import get_logger

logger = get_logger(__name__)

PEDIFORM_HEADERS = [
    "Patient First Name",
    "Patient Last Name",
    "Patient DOB",
    "Appointment",
    "Visit Type",
]


def _fmt_date(d: Optional[date]) -> str:
    """Format date as MM/DD/YYYY text string."""
    if d is None:
        return ""
    return d.strftime("%m/%d/%Y")


def _fmt_appointment(appt_date: Optional[date], appt_time) -> str:
    """Combine date + time into 'MM/DD/YYYY HH:MM' text."""
    if appt_date is None:
        return ""
    date_str = _fmt_date(appt_date)
    if appt_time is None:
        return date_str
    # appt_time may be datetime.time
    if isinstance(appt_time, dt_time):
        return f"{date_str} {appt_time.strftime('%H:%M')}"
    return date_str


def build_pediform_excel(
    patients: List[Dict],
    output_path: str = PEDIFORM_IMPORT_PATH,
) -> str:
    """
    Write a clean PediForm import Excel from the eligible patient list.

    Args:
        patients:    Output of excel_reader.get_eligible_patients()
        output_path: Where to save the file (default from settings)

    Returns:
        The path where the file was saved.
    """
    if not patients:
        logger.warning("No eligible patients — PediForm import file will be empty (headers only).")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Write headers
    ws.append(PEDIFORM_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Write patient rows — ALL values as text strings
    for p in patients:
        row = [
            p["first_name"],
            p["last_name"],
            _fmt_date(p["dob"]),
            _fmt_appointment(p["appt_date"], p.get("appt_time")),
            p.get("visit_type", "Well Visit"),
        ]
        ws.append(row)

    # Force text format on every data cell so Excel/PediForm can't
    # misinterpret the date strings as date serials
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.number_format = "@"   # "@" = text in Excel

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output_path)
    logger.info(
        f"PediForm import file saved: {output_path} "
        f"({len(patients)} patient(s))"
    )
    return output_path
