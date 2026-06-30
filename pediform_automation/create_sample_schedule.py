"""
Creates a sample daily_schedule.xlsx with dates stored as TEXT (not Excel date serials).
PediForm requires plain text dates — Excel's native date type sends a number instead.

Run once to generate the test file, then import it into PediForm.
"""

import openpyxl
from openpyxl.styles import Font

OUTPUT_PATH = r"C:\Users\prase\Downloads\automation_pd_forms\daily_schedule.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Schedule"

# Headers — must match PediForm exactly
headers = [
    "Patient First Name",
    "Patient Last Name",
    "Patient DOB",
    "Appointment",
    "Visit Type",
]
ws.append(headers)

# Make headers bold
for cell in ws[1]:
    cell.font = Font(bold=True)

# Sample patient row
# IMPORTANT: All values written as plain strings, NOT date objects.
# PediForm expects MM/DD/YYYY for dates.
ws.append([
    "Test",
    "Patient",
    "04/26/2026",       # Patient DOB as TEXT string (MM/DD/YYYY)
    "06/27/2026 08:30", # Appointment as TEXT string
    "Well visit",
])

# Force every cell in the data rows to be stored as text (not date/number)
from openpyxl.styles import numbers
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.number_format = "@"  # "@" = text format in Excel

wb.save(OUTPUT_PATH)
print(f"Saved: {OUTPUT_PATH}")
print("Now re-run pediform_import_test.py")
