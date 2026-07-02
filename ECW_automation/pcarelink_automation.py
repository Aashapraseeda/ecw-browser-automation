import asyncio
import os
import openpyxl
from dotenv import load_dotenv
from playwright.async_api import async_playwright

load_dotenv()

EXCEL_PATH = os.getenv("EXCEL_PATH")
EMAIL = os.getenv("PCARELINK_EMAIL")
PASSWORD = os.getenv("PCARELINK_PASSWORD")
PRACTICE = os.getenv("PCARELINK_PRACTICE")
MESSAGE = os.getenv("PCARELINK_MESSAGE")

def read_patients_from_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}
    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acct_no = row[col["Patient Acct No"]]
        full_name = row[col["Patient Name"]]
        if not acct_no:
            continue
        patients.append({
            "acct_no": str(acct_no).strip(),
            "name": str(full_name).strip() if full_name else "",
        })
    return patients

async def send_message_for_patient(page, patient):
    acct_no = patient["acct_no"]
    print(f"\nProcessing patient acct {acct_no} ({patient['name']})")

    # --- SEARCH PATIENT ---
    search_box = page.get_by_role("searchbox", name="Enter patient first name or")
    await search_box.click()
    await search_box.fill(acct_no)
    await page.wait_for_timeout(2000)

    # --- SELECT PATIENT FROM RESULTS ---
    try:
        await page.get_by_text(patient['name'].upper()).first.click(timeout=10000)
        print(f"Patient selected!")
    except:
        try:
            # Try clicking first result if exact name match fails
            await page.locator(".patient-result, .search-result").first.click(timeout=5000)
            print("Patient selected from first result!")
        except:
            print(f"Could not find patient {acct_no} - skipping")
            return

    await page.wait_for_timeout(1000)

    # --- CLICK SEND MESSAGE ---
    await page.locator('[data-test-id="pcl-payments-sendMessageLinkGuarantorDrawer"]').click()
    await page.wait_for_timeout(1000)
    print("Send message panel opened!")

    # --- SELECT PRACTICE ---
    try:
        await page.get_by_role("button", name=PRACTICE).click(timeout=5000)
        await page.wait_for_timeout(500)
        await page.get_by_role("button", name=PRACTICE).click(timeout=5000)
        await page.wait_for_timeout(500)
    except:
        print("Practice selection skipped")

    # --- SELECT APPOINTMENT SCHEDULING ---
    try:
        await page.get_by_role("menuitem", name="Appointment Scheduling").get_by_role("radio").check(timeout=5000)
        await page.wait_for_timeout(500)
        # Close dropdown
        await page.locator("#menu- > div").first.click(timeout=5000)
        await page.wait_for_timeout(500)
    except:
        print("Message type selection skipped")

    # --- TYPE MESSAGE ---
    message_box = page.get_by_role("textbox", name="Type your response and send")
    await message_box.click()
    await message_box.fill(MESSAGE)
    print("Message typed!")

    # --- SEND MESSAGE ---
    await page.locator('[data-test-id="pcl-payments-sendMessageButton"]').click()
    print(f"Message sent successfully!")
    await page.wait_for_timeout(1000)

    # --- CLOSE PATIENT ---
    await page.locator('[data-test-id="pcl-appointments-closePatientsDetails"]').click()
    await page.wait_for_timeout(1000)
    print("Patient closed!")

async def main():
    patients = read_patients_from_excel()
    print(f"Found {len(patients)} patients in Excel")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=300)
        context = await browser.new_context()
        page = await context.new_page()

        # --- LOGIN ---
        print("Logging into pcarelink...")
        await page.goto("https://app.pcarelink.com/login")
        await page.get_by_role("textbox", name="Enter email id").fill(EMAIL)
        await page.get_by_role("textbox", name="Enter password").fill(PASSWORD)
        await page.locator('[data-test-id="pcl-login-signInButton"]').click()
        await page.wait_for_load_state("networkidle")
        print("Logged in!")

        # --- NAVIGATE TO PATIENTS ---
        await page.locator('[data-test-id="pcl-menuDropDownComponent"]').click()
        await page.locator('[data-test-id="pcl-dashboard-popOver1"]').click()
        await page.wait_for_load_state("networkidle")
        print("Navigated to patients!")

        # --- FILTER BY PRACTICE ---
        await page.get_by_role("button", name="Filter by Practice").click()
        await page.get_by_text(f"{PRACTICE}Round Rock, us").click()
        await page.wait_for_timeout(2000)
        print(f"Filtered by practice: {PRACTICE}")

        # --- PROCESS EACH PATIENT ---
        for patient in patients:
            try:
                await send_message_for_patient(page, patient)
            except Exception as e:
                print(f"Error processing {patient['acct_no']}: {e}")
                try:
                    await page.locator('[data-test-id="pcl-appointments-closePatientsDetails"]').click()
                except:
                    pass
                continue

        print("\nAll patients processed!")
        await page.wait_for_timeout(60000)
        await browser.close()

if __name__ == "__main__":
    asyncio.run(main())