import asyncio
import os
import openpyxl
from dotenv import load_dotenv
from playwright.async_api import async_playwright

load_dotenv()

EXCEL_PATH = os.getenv("EXCEL_PATH")
ORG_NAME = os.getenv("PEDIFORMS_ORG")
EMAIL = os.getenv("PEDIFORMS_EMAIL")
PASSWORD = os.getenv("PEDIFORMS_PASSWORD")

VISIT_TYPE_TO_FORM = {
    "9 MONTH WC": "ASQ9Mos",
    "12 MONTH WC": "ASQ 12 Months",
    "1 YR WC": "ASQ 12 Months",
    "18 MONTH WC": "ASQ 18 Months",
    "24 MONTH WC": "ASQ 24 Months",
    "2 YR WC": "ASQ 24 Months",
    "30 MONTH WC": "ASQ30",
    "3 YEAR WC": "ASQ 36 Months",
    "36 MONTH WC": "ASQ 36 Months",
    "4 YEAR WC": "ASQ 48 Months",
    "48 MONTH WC": "ASQ 48 Months",
}

def read_patients_from_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}
    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acct_no = row[col["Patient Acct No"]]
        visit_type_raw = row[col["Visit Type"]]
        full_name = row[col["Patient Name"]]
        if not acct_no or not visit_type_raw:
            continue
        visit_type_desc = str(visit_type_raw).split(":")[-1].strip().upper()
        form_name = None
        for key, form in VISIT_TYPE_TO_FORM.items():
            if key.upper() == visit_type_desc:
                form_name = form
                break
        patients.append({
            "acct_no": str(acct_no).strip(),
            "name": str(full_name).strip() if full_name else "",
            "visit_type": visit_type_desc,
            "form_name": form_name,
        })
    return patients

async def go_back_to_patient_list(page):
    await page.get_by_role("link", name="← Back to today's patients").click()
    await page.wait_for_load_state("networkidle")

async def send_form_for_patient(page, patient):
    acct_no = patient["acct_no"]
    form_name = patient["form_name"]
    print(f"\nProcessing patient acct {acct_no} ({patient['name']}) - visit: {patient['visit_type']}")
    if not form_name:
        print(f"  No ASQ form mapped for '{patient['visit_type']}' - skipping")
        return
    search_box = page.get_by_role("textbox", name="Search…")
    await search_box.click()
    await search_box.fill(acct_no)
    await page.wait_for_timeout(1000)
    try:
        await page.get_by_role("link", name="View").click(timeout=10000)
    except Exception:
        print(f"  Could not find patient {acct_no} - skipping")
        return
    await page.get_by_role("button", name="+ Send a form").click()
    await page.wait_for_timeout(500)
    try:
        await page.locator("label").filter(has_text=form_name).first.click(timeout=10000)
    except Exception:
        print(f"  Could not find form '{form_name}' - skipping")
        await go_back_to_patient_list(page)
        return
    await page.get_by_role("button", name="Send form").click()
    print(f"  Sent '{form_name}' successfully!")
    await go_back_to_patient_list(page)

async def main():
    patients = read_patients_from_excel()
    print(f"Found {len(patients)} patients in Excel")
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=300)
        context = await browser.new_context()
        page = await context.new_page()
        print("Logging into Pediforms...")
        await page.goto("https://admin.pediformpro.com/staff/login")
        await page.get_by_role("textbox", name="Organization name").fill(ORG_NAME)
        await page.get_by_role("textbox", name="Email").fill(EMAIL)
        await page.get_by_role("textbox", name="Password").fill(PASSWORD)
        await page.get_by_role("button", name="Sign in").click()
        await page.wait_for_load_state("networkidle")
        print("Logged in!")
        print("Uploading schedule Excel...")
        await page.get_by_role("button", name="Choose File").set_input_files(EXCEL_PATH)
        await page.get_by_role("button", name="Import schedule").click()
        await page.wait_for_load_state("networkidle")
        print("Schedule imported!")
        for patient in patients:
            try:
                await send_form_for_patient(page, patient)
            except Exception as e:
                print(f"  Error: {e}")
                try:
                    await go_back_to_patient_list(page)
                except Exception:
                    pass
                continue
        print("\nAll patients processed!")
        await page.wait_for_timeout(60000)
        await browser.close()

if __name__ == "__main__":
    asyncio.run(main())