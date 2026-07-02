import asyncio
import os
import glob
import json
import openpyxl
from dotenv import load_dotenv
from playwright.async_api import async_playwright

load_dotenv()

USERNAME = os.getenv("ECW_USERNAME")
PASSWORD = os.getenv("ECW_PASSWORD")
EXCEL_PATH = os.getenv("EXCEL_PATH")
DOC_FOLDER = os.getenv("ECW_PATIENTS_DOC_FOLDER")

def read_patients():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}
    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        last_name = row[col["Patient Last Name"]]
        first_name = row[col["Patient First Name"]]
        acct_no = row[col["Patient Acct No"]]
        if not last_name or not first_name:
            continue
        patients.append({
            "first_name": str(first_name).strip(),
            "last_name": str(last_name).strip(),
            "acct_no": str(acct_no).strip() if acct_no else "",
            "search_name": f"{last_name},{first_name}".strip(),
            "folder_name": f"{last_name} {first_name}_doc".strip()
        })
    return patients

def get_all_files(patient):
    folder_path = os.path.join(DOC_FOLDER, patient["folder_name"])
    if not os.path.exists(folder_path):
        print(f"Folder not found: {folder_path}")
        return []
    files = glob.glob(os.path.join(folder_path, "*"))
    if not files:
        print(f"No files found in: {folder_path}")
        return []
    files.sort(key=os.path.getmtime)
    print(f"Found {len(files)} files in folder")
    return files

async def get_existing_docs(page):
    existing_docs = []
    try:
        doc_links = await page.locator('a[id^="patientdocsTreeLink"]').all()
        for link in doc_links:
            doc_object = await link.get_attribute('document-object')
            if doc_object:
                doc_data = json.loads(doc_object)
                label = doc_data.get('label', '').strip().lower()
                if label:
                    existing_docs.append(label)
        print(f"Existing docs in eCW: {existing_docs}")
    except Exception as e:
        print(f"Could not read existing docs: {e}")
    return existing_docs

async def go_to_patient_search(page):
    print("Closing modals...")
    await page.keyboard.press("Escape")
    await asyncio.sleep(1)
    await page.keyboard.press("Escape")
    await asyncio.sleep(1)
    try:
        await page.wait_for_selector("#patient-hubBtn1", timeout=10000)
        await page.locator("#patient-hubBtn1").click()
        await asyncio.sleep(1)
    except:
        pass
    await page.wait_for_selector("#jellybean-panelLink65", timeout=30000)
    await page.locator("#jellybean-panelLink65").click()
    await page.get_by_role("textbox", name="Last Name, First Name").wait_for(timeout=30000)
    print("Patient search ready!")

async def upload_patient_docs():
    patients = read_patients()
    print(f"Found {len(patients)} patients in Excel")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=200)
        context = await browser.new_context()
        page = await context.new_page()

        # --- LOGIN ---
        print("Opening eCW login page...")
        await page.goto("https://txsnmbapp.ecwcloud.com/mobiledoc/jsp/webemr/login/newLogin.jsp")
        await page.get_by_role("textbox", name="Enter username to continue").fill(USERNAME)
        await page.get_by_role("button", name="Next").click()
        await page.get_by_role("textbox", name="Enter Password to continue").fill(PASSWORD)
        await page.get_by_role("button", name="Log In").click()
        print("Login submitted...")

        # --- WAIT FOR HOME PAGE ---
        await page.wait_for_selector('#jellybean-panelLink33', timeout=120000)
        print("Home page loaded!")

        # --- WAIT FOR LOADING OVERLAY ---
        try:
            await page.wait_for_selector('#load', state='hidden', timeout=120000)
            print("ECW fully loaded!")
        except:
            print("Loading overlay already hidden!")

        # --- HANDLE LICENSE ALERT ---
        print("Checking for License Alert...")
        dismissed = False
        for i in range(20):
            try:
                if await page.locator("#providerLicense button.clsMyButton").is_visible():
                    await page.click("#providerLicense button.clsMyButton")
                    print(f"License Alert dismissed at second {i}!")
                    dismissed = True
                    break
            except:
                pass
            await asyncio.sleep(1)
        if not dismissed:
            print("No License Alert, continuing...")

        # --- OPEN PATIENT SEARCH FIRST TIME ---
        print("Opening Patient Search...")
        await page.wait_for_selector("#jellybean-panelLink65", timeout=30000)
        await page.locator("#jellybean-panelLink65").click()
        await page.get_by_role("textbox", name="Last Name, First Name").wait_for(timeout=30000)
        print("Patient search ready!")

        # --- PROCESS EACH PATIENT ---
        for index, patient in enumerate(patients):
            print(f"\nProcessing patient {index+1}/{len(patients)}: {patient['last_name']} {patient['first_name']}")

            all_files = get_all_files(patient)
            if not all_files:
                print(f"Skipping - no files found in folder: {patient['folder_name']}")
                continue

            try:
                # --- SEARCH PATIENT ---
                print(f"Searching for: {patient['search_name']}")
                search_box = page.get_by_role("textbox", name="Last Name, First Name")
                await search_box.wait_for(timeout=30000)
                await search_box.fill(patient['search_name'])
                await page.wait_for_selector("#patientLName1", timeout=30000)

                # --- SELECT PATIENT ---
                try:
                    await page.get_by_role("cell", name=patient['last_name'], exact=False).first.click()
                    await page.get_by_text(patient['last_name'], exact=False).first.click()
                    print("Patient selected!")
                except:
                    print(f"Could not find patient - skipping")
                    await go_to_patient_search(page)
                    continue

                # Handle "Please select a patient" popup
                try:
                    if await page.locator("text=Please select a patient").is_visible():
                        await page.get_by_role("button", name="OK").click()
                        print("Handled patient selection popup!")
                        await go_to_patient_search(page)
                        continue
                except:
                    pass

                # --- GO TO PATIENT DOCS ---
                print("Opening Patient Docs...")
                await page.wait_for_selector('button:has-text("Patient Docs")', timeout=30000)
                await page.get_by_role("button", name="Patient Docs").click()

                # --- SEARCH FOR CHART DOCUMENTS ---
                print("Searching for Chart Documents...")
                await page.get_by_role("textbox", name="Quick Search").wait_for(timeout=30000)
                await page.get_by_role("textbox", name="Quick Search").fill("chart")
                await page.wait_for_selector('a:has-text("Chart Documents")', timeout=30000)
                await page.locator("a").filter(has_text="Chart Documents").nth(1).click()
                await asyncio.sleep(2)

                # --- READ EXISTING DOCS ---
                print("Reading existing documents...")
                existing_docs = await get_existing_docs(page)

                # --- FILTER OUT ALREADY UPLOADED FILES ---
                new_files = []
                for f in all_files:
                    filename = os.path.splitext(os.path.basename(f))[0].lower()
                    if filename in existing_docs:
                        print(f"Skipping already uploaded: {filename}")
                    else:
                        new_files.append(f)

                if not new_files:
                    print("All files already uploaded - skipping patient!")
                    await go_to_patient_search(page)
                    continue

                print(f"{len(new_files)} new files to upload")

                # --- UPLOAD NEW FILES ONE BY ONE ---
                for file_index, file_path in enumerate(new_files):
                    print(f"Uploading file {file_index+1}/{len(new_files)}: {file_path}")

                    if file_index > 0:
                        await page.get_by_role("textbox", name="Quick Search").wait_for(timeout=30000)
                        await page.get_by_role("textbox", name="Quick Search").fill("chart")
                        await page.wait_for_selector('a:has-text("Chart Documents")', timeout=30000)
                        await page.locator("a").filter(has_text="Chart Documents").nth(1).click()
                        await asyncio.sleep(1)

                    async with page.expect_file_chooser() as fc_info:
                        await page.wait_for_selector("#patientdocsBtn4", timeout=30000)
                        await page.locator("#patientdocsBtn4").click()
                    file_chooser = await fc_info.value
                    await file_chooser.set_files(file_path)
                    await asyncio.sleep(1)
                    print("File selected!")

                    try:
                        if await page.locator("text=Please select a category").is_visible():
                            await page.get_by_role("button", name="OK").click()
                            print("Category popup dismissed!")
                    except:
                        pass

                    await page.wait_for_selector('button.commonButton:has-text("OK")', timeout=30000)
                    await page.locator('button.commonButton:has-text("OK")').click()
                    print("Flatten PDF dismissed!")

                    await page.wait_for_selector('#btnOk', timeout=30000)
                    await page.locator('#btnOk').click()
                    print(f"File {file_index+1} saved successfully!")
                    await asyncio.sleep(1)

                print(f"All new files uploaded for {patient['last_name']} {patient['first_name']}!")

                # --- CLOSE PATIENT AND GO BACK TO SEARCH ---
                print("Closing patient and going back to search...")
                await go_to_patient_search(page)
                print("Ready for next patient!")

            except Exception as e:
                print(f"Error: {e}")
                try:
                    await go_to_patient_search(page)
                except:
                    pass
                continue

        print("\nAll patients processed successfully!")
        await asyncio.sleep(99999)

asyncio.run(upload_patient_docs())