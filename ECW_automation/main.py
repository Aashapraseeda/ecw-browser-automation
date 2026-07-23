import asyncio
import os
import re
import glob
import json
import openpyxl
from datetime import date, datetime, timedelta
from dotenv import load_dotenv
from playwright.async_api import async_playwright

import state_db

load_dotenv()

# --- CREDENTIALS ---
ECW_USERNAME = os.getenv("ECW_USERNAME")
ECW_PASSWORD = os.getenv("ECW_PASSWORD")
ECW_LOGIN_URL = os.getenv("ECW_LOGIN_URL", "https://txsnmbapp.ecwcloud.com/mobiledoc/jsp/webemr/login/newLogin.jsp")
ECW_EBO_HOME_URL = os.getenv("ECW_EBO_HOME_URL", "https://txsnmbebo.ecwcloud.com/bi/?perspective=home")
PEDIFORMS_ORG = os.getenv("PEDIFORMS_ORG")
PEDIFORMS_EMAIL = os.getenv("PEDIFORMS_EMAIL")
PEDIFORMS_PASSWORD = os.getenv("PEDIFORMS_PASSWORD")
PCARELINK_EMAIL = os.getenv("PCARELINK_EMAIL")
PCARELINK_PASSWORD = os.getenv("PCARELINK_PASSWORD")
PCARELINK_PRACTICE = os.getenv("PCARELINK_PRACTICE")
PCARELINK_MESSAGE = os.getenv("PCARELINK_MESSAGE")
EXCEL_PATH = os.getenv("EXCEL_PATH")
FILTERED_EXCEL_PATH = os.path.join(os.path.dirname(EXCEL_PATH), "filtered_schedule.xlsx") if EXCEL_PATH else None
DOC_FOLDER = os.getenv("ECW_PATIENTS_DOC_FOLDER")

# --- SETTINGS ---
# NOTE: the completed-form check is now a single pass per run - cron itself
# (running this script every few hours) provides the "check again later"
# behavior. There is no internal wait loop anymore.
#
# Reminder messages are intentionally not implemented yet - to be added
# separately once the cron architecture is finalized.

STATE_RETENTION_DAYS = int(os.getenv("STATE_RETENTION_DAYS", "30"))

VISIT_TYPE_TO_FORM = {
    "9 MONTH WC": "ASQ9Mos",
    "12 MONTH WC": "ASQ 12 Months",
    "12 MONTHWC": "ASQ 12 Months",
    "1 YEAR WC": "ASQ 12 Months",
    "15 MONTH WC": "ASQ 18 Months",
    "15 MONTHWC": "ASQ 18 Months",
    "18 MONTH WC": "ASQ 18 Months",
    "18 MONTHWC": "ASQ 18 Months",
    "24 MONTH WC": "ASQ 24 Months",
    "24 MONTHWC": "ASQ 24 Months",
    "2 YEAR WC": "ASQ 24 Months",
    "30 MONTH WC": "ASQ30",
    "30 MONTHWC": "ASQ30",
    "3 YEAR WC": "ASQ 36 Months",
    "36 MONTH WC": "ASQ 36 Months",
    "36 MONTHWC": "ASQ 36 Months",
    "4 YEAR WC": "ASQ 48 Months",
    "48 MONTH WC": "ASQ 48 Months",
    "48 MONTHWC": "ASQ 48 Months",
}

VISIT_TYPE_TO_FORM_FILENAME = {
    "9 MONTH WC": "ASQ9Mos",
    "12 MONTH WC": "ASQ_12_Months",
    "12 MONTHWC": "ASQ_12_Months",
    "1 YEAR WC": "ASQ_12_Months",
    "15 MONTH WC": "ASQ_18_Months",
    "15 MONTHWC": "ASQ_18_Months",
    "18 MONTH WC": "ASQ_18_Months",
    "18 MONTHWC": "ASQ_18_Months",
    "24 MONTH WC": "ASQ_24_Months",
    "24 MONTHWC": "ASQ_24_Months",
    "2 YEAR WC": "ASQ_24_Months",
    "30 MONTH WC": "ASQ30",
    "30 MONTHWC": "ASQ30",
    "3 YEAR WC": "ASQ_36_Months",
    "36 MONTH WC": "ASQ_36_Months",
    "36 MONTHWC": "ASQ_36_Months",
    "4 YEAR WC": "ASQ_48_Months",
    "48 MONTH WC": "ASQ_48_Months",
    "48 MONTHWC": "ASQ_48_Months",
}

# --- M-CHAT + TB (2026-07-22 addition) ---
# M-CHAT is sent ALONGSIDE the age-appropriate ASQ form (never instead of
# it) for these visit-type KEYS only - NOT gated on the resolved form
# NAME, because VISIT_TYPE_TO_FORM deliberately maps BOTH "15 MONTH WC"
# and "18 MONTH WC" to the same form text ("ASQ 18 Months" - the 15-month
# bracket reuses the 18-month form) - gating on form name would wrongly
# also trigger M-CHAT for genuine 15-month visits.
MCHAT_TRIGGER_VISIT_TYPES = {"18 MONTH WC", "18 MONTHWC", "24 MONTH WC", "24 MONTHWC", "2 YEAR WC"}
MCHAT_FORM = {"form_name": "M-Chat", "form_filename": "M_Chat"}

# TB form: every Well Check patient aged 12 months - 18 years inclusive
# gets a TB form, INDEPENDENT of ASQ/M-CHAT eligibility - e.g. a 5-year-old
# Well Check gets TB only (no ASQ form exists for that age, so previously
# such patients were excluded entirely - see _forms_for_patient below).
TB_MIN_AGE_MONTHS = 12
TB_MAX_AGE_MONTHS = 18 * 12  # 216 - "18 years inclusive"
TB_FORM = {"form_name": "TB", "form_filename": "TB"}

# "TB" and "M-Chat" (mixed case, not "M-CHAT") were CONFIRMED LIVE
# (2026-07-22, read-only inspection, no forms sent) against Lone Star's
# Patient Forms Now account. This project's OWN Pediforms account
# (admin.pediformpro.com) could NOT be independently checked in the same
# session - the demo test patient couldn't be located due to a pre-existing,
# unrelated week-filter/date-rollover issue (blocks chart access, not new).
# Applying the same label text here is a reasonable INFERENCE (identical
# UI text/flow observed everywhere else across both accounts - same
# "+ Send a form", "Send form", "← Back to today's patients" wording -
# strongly suggests the same underlying product), but is NOT independently
# confirmed for this specific account. If the real text differs, the send
# step's has_text=... filter will simply not find a match, log "Form not
# found - skipping", and move on without crashing - but that form will
# silently never be sent until confirmed and corrected here.
#
# Multi-select: CONFIRMED on Lone Star's account that the checkbox panel
# supports checking several forms before one combined Send click (see
# lone_star_automation/config/settings.py for detail). pediforms_send_forms()
# below now checks every applicable box once and sends once, applying the
# same inference to this account for the same reason as above.


def _parse_date_flexible(value):
    """
    Local port of lone_star_automation/utils/date_utils.py's
    parse_date_flexible() - this project has no shared date-utils module,
    so this is intentionally duplicated rather than cross-imported from
    the separate Lone Star project.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%m/%d/%Y %H:%M", "%m/%d/%Y", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _age_in_months(dob, reference_date):
    """Local port of lone_star_automation's age_in_months() - see note above."""
    months = (reference_date.year - dob.year) * 12 + (reference_date.month - dob.month)
    if reference_date.day < dob.day:
        months -= 1
    return months


async def _wait_for_loading_overlay_gone(page, timeout_ms=60000, retries=3):
    """
    (2026-07-23 fix) Robust wait for eCW's '#load' ("Building your user
    experience") overlay to be hidden. A single wait_for_selector(state=
    'hidden') call can resolve or except prematurely if the overlay
    toggles visibility multiple times during a complex page load - a live
    production run showed the overlay still intercepting pointer events on
    the very next click, well after the code had already logged "Loading
    screen already hidden!" (a bare except was silently assuming that
    meant done, which the same run disproved). Retries a few times rather
    than trusting one shot / one silent except.
    """
    for attempt in range(retries):
        try:
            await page.wait_for_selector('#load', state='hidden', timeout=timeout_ms)
            print("Loading overlay confirmed hidden.")
            return
        except Exception:
            print(f"Loading overlay still present or check unstable (attempt {attempt + 1}/{retries}) - re-checking...")
            await asyncio.sleep(2)
    print("Proceeding despite loading-overlay uncertainty after retries.")


def _forms_for_patient(visit_type_desc, age_months):
    """
    Returns the list of {"form_name","form_filename"} dicts this Well
    Check patient should receive - three independent rules:
      1. ASQ - existing VISIT_TYPE_TO_FORM/VISIT_TYPE_TO_FORM_FILENAME
         text-based lookup, UNCHANGED (still the sole source of ASQ
         eligibility/form choice - preserves the existing, proven rule).
      2. M-CHAT - additional, sent alongside ASQ only when the resolved
         ASQ form is the 18 or 24 month one (never instead of ASQ).
      3. TB - independent age test (12-216 months inclusive, from DOB) -
         applies even when no ASQ form matched (e.g. a 5-year-old Well
         Check, previously excluded entirely since VISIT_TYPE_TO_FORM has
         no entry for it).
    Returns [] if none apply - caller should treat that as "not eligible",
    same as before.
    """
    forms = []
    form_name = VISIT_TYPE_TO_FORM.get(visit_type_desc)
    if form_name:
        forms.append({
            "form_name": form_name,
            "form_filename": VISIT_TYPE_TO_FORM_FILENAME.get(visit_type_desc, "form"),
        })
        if visit_type_desc in MCHAT_TRIGGER_VISIT_TYPES:
            forms.append(dict(MCHAT_FORM))
    if age_months is not None and TB_MIN_AGE_MONTHS <= age_months <= TB_MAX_AGE_MONTHS:
        forms.append(dict(TB_FORM))
    return forms

# Appointments at these facilities are excluded from this clinic's filtered
# schedule - Lone Star Pediatrics Midlothian is handled by its own separate
# automation project now, so it must not be double-processed here. Compared
# case-insensitively with leading/trailing whitespace stripped.
EXCLUDED_FACILITY_NAMES = {"lone star pediatrics midlothian"}

# (2026-07-21) Maps eCW's raw "Appointment Facility Name" values to the
# practice name shown in ReachMyDr/PCareLink's "Filter by Practice"
# dropdown. One shared PCareLink account (aasha@painmedpa.com) covers all
# of these practices, so the correct filter now depends on each PATIENT's
# own facility, not a single fixed value. Keys are normalized (stripped +
# lowercased) before lookup.
#
# Exact string matches between the real downloaded Excel and a live,
# read-only ReachMyDr dropdown check, plus one explicitly-confirmed
# ambiguous case:
#   'Peds Center of Round Rock PA' had no exact match - two distinct,
#   separately-clickable ReachMyDr entries could apply ("Pediatric Center
#   Of Round Rock" vs "Ped Center Of Round Rock"); confirmed by the
#   clinic to be "Ped Center Of Round Rock".
#
# DELIBERATELY LEFT UNMAPPED (not ambiguous, just not set up in ReachMyDr
# yet - do not guess):
#   'Lone Star Pediatrics Midlothian' - matches ZERO entries in the
#   ReachMyDr dropdown (confirmed live). Not applicable to this project
#   anyway (excluded via EXCLUDED_FACILITY_NAMES above), relevant only to
#   the Lone Star project.
FACILITY_TO_PRACTICE = {
    "river ridge pediatrics": "River Ridge Pediatrics",
    "peds center of round rock pa": "Ped Center Of Round Rock",
    "elgin pediatrics": "Elgin Pediatrics",
    "pediatric care of austin": "Pediatric Care of Austin",
    "pediatric center of north austin": "Pediatric Center of North Austin",
}


def resolve_practice_for_facility(facility_name):
    """
    Returns the ReachMyDr practice name for a given eCW facility value, or
    None if unmapped. Callers must log a warning and SKIP the message when
    this returns None - never fall back to a default/guessed practice.
    """
    if not facility_name:
        return None
    return FACILITY_TO_PRACTICE.get(str(facility_name).strip().lower())

# ─────────────────────────────────────────
# SHARED HELPERS
# ─────────────────────────────────────────

def read_patients_from_excel():
    """
    Production filter: Well Check visits get ASQ (9-48mo, VISIT_TYPE_TO_FORM
    text-based, unchanged), plus M-CHAT (18/24mo ASQ only) and/or TB
    (12mo-18yr, DOB-based, independent of ASQ) - see _forms_for_patient().
    A patient is included if EITHER an ASQ form OR TB applies (a 5+ year
    Well Check with no ASQ form still qualifies for TB alone).
    No visit reason filter — all eligible patients included.
    Excludes EXCLUDED_FACILITY_NAMES (see above) before any other filtering.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}
    patients = []
    filtered_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acct_no = row[col["Patient Acct No"]]
        last_name = row[col["Patient Last Name"]]
        first_name = row[col["Patient First Name"]]
        visit_type_raw = row[col["Visit Type"]]
        appointment_date_raw = row[col["Appointment Date"]]
        dob_raw = row[col["Patient DOB"]] if "Patient DOB" in col else None
        if not acct_no:
            continue
        facility_name_raw = row[col["Appointment Facility Name"]] if "Appointment Facility Name" in col else None
        if facility_name_raw and str(facility_name_raw).strip().lower() in EXCLUDED_FACILITY_NAMES:
            print(f"Skipping {last_name} {first_name} - excluded facility: {facility_name_raw!r}")
            continue

        visit_type_desc = str(visit_type_raw).split(":")[-1].strip().upper() if visit_type_raw else ""
        is_well_check = visit_type_desc.endswith(" WC") or visit_type_desc == "WC"
        if not is_well_check:
            print(f"Skipping {last_name} {first_name} - not a Well Check visit type: {visit_type_desc!r}")
            continue

        dob = _parse_date_flexible(dob_raw)
        appt_date = _parse_date_flexible(appointment_date_raw)
        age_months = _age_in_months(dob, appt_date) if dob and appt_date else None

        forms = _forms_for_patient(visit_type_desc, age_months)
        if not forms:
            print(f"Skipping {last_name} {first_name} - no ASQ/M-CHAT/TB form applies "
                  f"(visit type {visit_type_desc!r}, age {age_months} month(s))")
            continue

        patients.append({
            "acct_no": str(acct_no).strip(),
            "appointment_date": state_db.normalize_date(appointment_date_raw),
            "last_name": str(last_name).strip() if last_name else "",
            "first_name": str(first_name).strip() if first_name else "",
            "folder_name": f"{last_name} {first_name}_doc".strip(),
            "search_name": f"{last_name},{first_name}".strip(),
            "visit_type": visit_type_desc,
            "forms": forms,
            "form_name": ", ".join(f["form_name"] for f in forms),
            "form_filename": "_".join(f["form_filename"] for f in forms),
            "facility": str(facility_name_raw).strip() if facility_name_raw else "",
        })
        filtered_rows.append(row)

    # Save filtered Excel for Pediforms import
    filtered_wb = openpyxl.Workbook()
    filtered_ws = filtered_wb.active
    filtered_ws.append(headers)
    for row in filtered_rows:
        filtered_ws.append(list(row))
    filtered_wb.save(FILTERED_EXCEL_PATH)
    print(f"Filtered schedule ({len(filtered_rows)} patients) saved to: {FILTERED_EXCEL_PATH}")
    return patients

def ensure_patient_folder(patient):
    folder_path = os.path.join(DOC_FOLDER, patient["folder_name"])
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Created folder: {folder_path}")
    return folder_path

# ─────────────────────────────────────────
# STEP 0 — ECW: EXPORT SCHEDULE
# ─────────────────────────────────────────

async def _wait_for_report_running_modal_gone(page1, max_checks=90):
    """
    Waits out eCW's "Your report is running" modal (a leftover/queued
    report execution, from this account's own prior report requests,
    still processing server-side). This modal sits on top of the whole
    report screen and blocks every click underneath it - including the
    date picker - which can look exactly like an unrelated UI bug if you
    only check for it once, early, and it appears again later.
    """
    try:
        if await page1.get_by_text("Your report is running").is_visible():
            print("Report already running - waiting...")
            for i in range(max_checks):
                if not await page1.get_by_text("Your report is running").is_visible():
                    print("Report finished.")
                    return
                print(f"Still running... ({i + 1}/{max_checks})")
                await asyncio.sleep(2)
            print("Report still running after max wait - proceeding anyway.")
    except:
        pass


async def click_calendar_option(iframe, day_str, descriptions_to_try, label, retries=3):
    """
    Try clicking calendar option with multiple description variants then
    fallback.

    (2026-07-24) Live diagnostics on Lone Star's identical code confirmed
    the day option itself is genuinely visible, enabled, and has a real
    bounding box when this fails - the click is blocked by something
    transiently overlapping it, not a missing/broken element. Same
    pattern already solved elsewhere in this codebase family (the
    Facility tab click in lone_star_automation/ecw/facility_filter.py) -
    force=True bypasses Playwright's actionability/overlap check and
    clicks the element's coordinates directly. Outer retry/wait loop kept
    as a second layer in case the widget is still mid-render.
    """
    last_error = None
    for attempt in range(retries):
        for desc in descriptions_to_try:
            try:
                await iframe.get_by_role("option", name=day_str, description=desc, exact=True).click(timeout=5000, force=True)
                print(f"{label} date set! (description='{desc}')")
                return
            except:
                pass
        # Final fallback - no description
        try:
            await iframe.get_by_role("option", name=day_str, exact=True).first.click(timeout=5000, force=True)
            print(f"{label} date set! (no description fallback)")
            return
        except Exception as e:
            last_error = e
            print(f"Could not click {label} date option (attempt {attempt + 1}/{retries}) - retrying...")
            await asyncio.sleep(2)
    raise RuntimeError(f"Could not set {label} date to day {day_str} after {retries} attempts: {last_error}")

async def ecw_export_schedule():
    print("\n" + "="*50)
    print("STEP 0 — ECW: EXPORTING SCHEDULE (TOMORROW THROUGH +3 DAYS)")
    print("="*50)

    # Both 2-letter and 3-letter day description variants
    day_desc_2 = {0: "Mo", 1: "Tu", 2: "We", 3: "Th", 4: "Fr", 5: "Sa", 6: "Su"}
    day_desc_3 = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}

    # (2026-07-21) Changed from a 7-day window starting today to a 3-day
    # window starting tomorrow, per explicit request - report should never
    # include today's appointments, only tomorrow through +3 days.
    start_date = date.today() + timedelta(days=1)
    end_date = date.today() + timedelta(days=3)
    start_day_str = str(start_date.day)
    end_day_str = str(end_date.day)
    start_descs = [day_desc_2[start_date.weekday()], day_desc_3[start_date.weekday()]]
    end_descs = [day_desc_2[end_date.weekday()], day_desc_3[end_date.weekday()]]
    print(f"Date range: {start_date} to {end_date}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()
        page = await context.new_page()

        # --- LOGIN ---
        print("Logging into eCW...")
        await page.goto(
            ECW_LOGIN_URL,
            timeout=120000, wait_until="domcontentloaded"
        )
        await asyncio.sleep(3)
        await page.get_by_role("textbox", name="Enter username to continue").fill(ECW_USERNAME)
        await page.get_by_role("button", name="Next").click()
        await asyncio.sleep(8)
        await page.click('input[type="password"]')
        await page.keyboard.type(ECW_PASSWORD)
        await page.keyboard.press("Enter")
        print("Login submitted...")

        await page.wait_for_selector('#jellybean-panelLink33', timeout=120000)
        print("Home page loaded!")

        print("Waiting for eCW to fully load...")
        await _wait_for_loading_overlay_gone(page)

        print("Checking for License Alert...")
        dismissed = False
        for i in range(20):
            try:
                if await page.locator("#providerLicense button.clsMyButton").is_visible():
                    await page.click("#providerLicense button.clsMyButton")
                    print("License Alert dismissed!")
                    dismissed = True
                    break
            except:
                pass
            await asyncio.sleep(1)
        if not dismissed:
            print("No License Alert, continuing...")
        await asyncio.sleep(2)

        # Re-verify right before the next click - the overlay can
        # re-render asynchronously after the first check passes (see
        # _wait_for_loading_overlay_gone's docstring).
        await _wait_for_loading_overlay_gone(page)

        # --- NAVIGATE TO EBO REPORTS ---
        # (2026-07-24) force=True on these: live runs showed different
        # transient backdrop elements (e.g. #pnBackDrop, seen even after
        # the loading overlay was already confirmed hidden) intermittently
        # blocking these clicks - a different element each time, not the
        # same bug recurring. Same fix already proven for the calendar day
        # click and the Facility tab click on Lone Star's identical code.
        print("Navigating to eBO Reports...")
        await page.locator("#jellybean-panelLink4").click(force=True)
        await asyncio.sleep(1)
        await page.get_by_text("Menu", exact=True).click(force=True)
        await asyncio.sleep(1)
        await page.locator("#pane6").get_by_text("Reports").click(force=True)
        await asyncio.sleep(1)

        async with page.expect_popup(timeout=60000) as page1_info:
            await page.get_by_text("eBO Reports CTRL + ALT + E").click()
        page1 = await page1_info.value
        await page1.goto(
            ECW_EBO_HOME_URL,
            timeout=60000, wait_until="domcontentloaded"
        )
        await asyncio.sleep(5)
        print("eBO Reports opened!")

        # --- NAVIGATE TO ENCOUNTER PATIENT DOWNLOAD ---
        await page1.get_by_role("link", name="eCWEBO", exact=True).click()
        await asyncio.sleep(2)
        await page1.get_by_role("link", name="- Administrative Reports").click()
        await asyncio.sleep(2)
        await page1.get_by_role("link", name="- Encounter Patient Download").click()
        await asyncio.sleep(8)
        print("Encounter Patient Download opened!")

        # --- WAIT IF REPORT ALREADY RUNNING ---
        await _wait_for_report_running_modal_gone(page1)

        # --- WAIT FOR IFRAME ---
        print("Waiting for iframe to load...")
        iframe = page1.locator("iframe[name=\"iD6D96C5E47F347C9B95828AC68A2D69B\"]").content_frame
        await iframe.get_by_role("img").first.wait_for(timeout=120000)

        # (2026-07-24 fix) A live diagnostic on Lone Star's identical
        # report screen showed the iframe's own content is unstable right
        # after this point - images can appear then disappear again
        # moments later as the report prompt panel (tabs: Additional
        # Prompts / Facility / Provider / Payer / Patient / Others) keeps
        # re-rendering. The date-range controls need TWO calendar icons
        # (img().first for start, img().nth(1) for end) - a live run
        # showed the count stabilizing at just 1, which the previous
        # version of this check accepted as "done" since it only required
        # non-zero, not the actual expected count. Now waits up to 90s
        # (was 30) and requires the count to reach >= 2 before considering
        # it settled - same stabilization pattern already proven for the
        # Facility results list in lone_star_automation/ecw/facility_filter.py,
        # now also checking for completeness, not just stability.
        previous_count = -1
        stable_checks = 0
        target_reached = False
        for _ in range(90):
            count = await iframe.get_by_role("img").count()
            if count >= 2 and count == previous_count:
                stable_checks += 1
                if stable_checks >= 3:
                    target_reached = True
                    break
            else:
                stable_checks = 0
            previous_count = count
            await asyncio.sleep(1)
        if target_reached:
            print(f"Iframe content stabilized at {previous_count} image(s).")
        else:
            print(f"Iframe content never reached 2+ stable images after extended wait "
                  f"(last count: {previous_count}) - proceeding anyway.")

        # (2026-07-24 fix) The "Your report is running" modal was observed
        # live (on Lone Star's identical code) appearing AFTER the check
        # above already passed - a report from an earlier attempt still
        # executing/queued server-side, only surfacing once the page
        # caught up. That modal sits on top of the date picker and blocks
        # every click under it (including the day option), which looked
        # like a calendar bug but wasn't - re-check right before dates.
        await _wait_for_report_running_modal_gone(page1)

        # --- SET DATES ---
        print(f"Setting start date: {start_date}")
        await iframe.get_by_role("img").first.click()
        await asyncio.sleep(2)
        await click_calendar_option(iframe, start_day_str, start_descs, "Start")
        await asyncio.sleep(1)

        print(f"Setting end date: {end_date}")
        await iframe.get_by_role("img").nth(1).click()
        await asyncio.sleep(2)
        await click_calendar_option(iframe, end_day_str, end_descs, "End")
        await asyncio.sleep(1)

        # --- CLICK OK ---
        await iframe.get_by_role("button", name="OK").click()
        await asyncio.sleep(2)
        print(f"Date range confirmed: {start_date} to {end_date}")

        # --- WAIT FOR REPORT TO GENERATE ---
        print("Waiting for report to generate (2-3 minutes)...")
        for i in range(120):
            try:
                is_disabled = await page1.locator("button[aria-label='Select a format']").get_attribute("disabled")
                if is_disabled is None:
                    print("Report ready!")
                    break
                print(f"Report still running... ({i+1}/120)")
            except:
                pass
            await asyncio.sleep(2)

        await asyncio.sleep(1)

        # --- DOWNLOAD EXCEL ---
        print("Clicking Select a format...")
        await page1.get_by_role("button", name="Select a format").click()
        await asyncio.sleep(3)
        await page1.get_by_role("link", name="Excel data").wait_for(timeout=15000)

        print("Clicking Excel data — waiting for download (may take 2-3 minutes)...")
        async with page1.expect_download(timeout=300000) as download_info:
            await page1.get_by_role("link", name="Excel data").click()
        download = await download_info.value
        await download.save_as(EXCEL_PATH)
        print(f"Excel saved to: {EXCEL_PATH}")

        await browser.close()
        print("\nSchedule exported successfully!")

# ─────────────────────────────────────────
# STEP 1 — PEDIFORMS: SEND FORMS
# ─────────────────────────────────────────

async def pediforms_send_forms(patients):
    print("\n" + "="*50)
    print("STEP 1 — PEDIFORMS: SENDING FORMS")
    print("="*50)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=300)
        context = await browser.new_context()
        page = await context.new_page()

        print("Logging into Pediforms...")
        await page.goto("https://admin.pediformpro.com/staff/login", timeout=60000, wait_until="domcontentloaded")
        await asyncio.sleep(3)
        await page.get_by_role("textbox", name="Organization name").fill(PEDIFORMS_ORG)
        await page.get_by_role("textbox", name="Email").fill(PEDIFORMS_EMAIL)
        await page.get_by_role("textbox", name="Password").fill(PEDIFORMS_PASSWORD)
        await page.get_by_role("button", name="Sign in").click()
        await page.wait_for_load_state("networkidle")
        print("Logged in!")

        print("Uploading filtered schedule Excel...")
        await page.get_by_role("button", name="Choose File").set_input_files(FILTERED_EXCEL_PATH)
        await page.get_by_role("button", name="Import schedule").click()
        await page.wait_for_load_state("networkidle")
        print("Schedule imported!")

        await page.get_by_role("combobox").nth(4).select_option("week")
        await page.wait_for_timeout(1000)

        for patient in patients:
            try:
                forms_to_send = patient.get("forms") or (
                    [{"form_name": patient["form_name"], "form_filename": patient["form_filename"]}]
                    if patient.get("form_name") else []
                )
                form_names = [f["form_name"] for f in forms_to_send]
                print(f"\nSending {len(forms_to_send)} form(s) for {patient['acct_no']} "
                      f"({patient['last_name']} {patient['first_name']}): {form_names}")
                search_box = page.get_by_role("textbox", name="Search…")
                await search_box.click()
                await search_box.fill(patient["acct_no"])
                await page.wait_for_timeout(1000)
                try:
                    await page.get_by_role("link", name="View").click(timeout=10000)
                except:
                    print(f"Patient {patient['acct_no']} not found - skipping")
                    continue

                # (2026-07-22) sends EVERY form for this patient (ASQ +
                # M-CHAT + TB, any combination) in ONE combined submission -
                # opens "+ Send a form" once, checks every matching box,
                # then clicks "Send form" once. Confirmed live on Lone
                # Star's account that this checkbox panel supports true
                # multi-select (see MCHAT_FORM's comment above) - applying
                # the same flow here as a reasonable inference for this
                # account, not independently confirmed (see same comment).
                checked_forms = []
                try:
                    await page.get_by_role("button", name="+ Send a form").click()
                    await page.wait_for_timeout(500)

                    for form in forms_to_send:
                        try:
                            await page.locator("label").filter(has_text=form["form_name"]).first.click(timeout=10000)
                            checked_forms.append(form)
                        except:
                            print(f"Form '{form['form_name']}' not found - skipping")
                            continue

                    if checked_forms:
                        await page.get_by_role("button", name="Send form").click()
                        print(f"Sent {[f['form_name'] for f in checked_forms]} successfully!")
                    else:
                        print(f"No matching form checkboxes found for {patient['acct_no']} - nothing sent")
                except Exception as e:
                    print(f"Error sending forms: {e}")

                await page.get_by_role("link", name="← Back to today's patients").click()
                await page.wait_for_load_state("networkidle")
            except Exception as e:
                print(f"Error: {e}")
                try:
                    await page.get_by_role("link", name="← Back to today's patients").click()
                    await page.wait_for_load_state("networkidle")
                except:
                    pass
                continue

        await browser.close()
        print("\nForms sent!")

# ─────────────────────────────────────────
# STEP 2 — PCARELINK: SEND MESSAGES
# ─────────────────────────────────────────

async def pcarelink_send_messages(patients):
    print("\n" + "="*50)
    print("STEP 2 — PCARELINK: SENDING MESSAGES")
    print("="*50)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=300)
        context = await browser.new_context()
        page = await context.new_page()

        print("Logging into pcarelink...")
        await page.goto("https://app.pcarelink.com/login", timeout=60000, wait_until="domcontentloaded")
        await asyncio.sleep(3)
        await page.get_by_role("textbox", name="Enter email id").fill(PCARELINK_EMAIL)
        await page.get_by_role("textbox", name="Enter password").fill(PCARELINK_PASSWORD)
        await page.locator('[data-test-id="pcl-login-signInButton"]').click()
        await asyncio.sleep(5)
        print("Logged in!")

        await page.locator('[data-test-id="pcl-menuDropDownComponent"]').click()
        await page.locator('[data-test-id="pcl-dashboard-popOver1"]').click()
        await page.wait_for_load_state("networkidle")

        for patient in patients:
            try:
                practice = resolve_practice_for_facility(patient.get("facility"))

                # --- TEMPORARY DEBUG LOGGING (per explicit request - remove once verified live) ---
                print(f"[DEBUG] Patient: {patient['first_name']} {patient['last_name']} | "
                      f"Facility: {patient.get('facility')!r} | Practice: {practice!r}")

                if not practice:
                    print(f"WARNING: no ReachMyDr practice mapping for facility {patient.get('facility')!r} "
                          f"(acct {patient['acct_no']}) - skipping message, NOT guessing a practice.")
                    continue

                print(f"\nSending message for {patient['acct_no']} ({patient['last_name']} {patient['first_name']})")
                await page.get_by_role("button", name="Filter by Practice").click()
                await page.get_by_text(practice, exact=False).click()
                await page.wait_for_timeout(2000)
                print(f"Filtered by practice: {practice}")

                search_box = page.get_by_role("searchbox", name="Enter patient first name or")
                await search_box.click()
                await search_box.fill(patient["acct_no"])
                await page.wait_for_timeout(2000)
                try:
                    await page.get_by_text(f"{patient['last_name'].upper()}, {patient['first_name'].upper()}").first.click(timeout=10000)
                except:
                    try:
                        await page.locator(".patient-result, .search-result").first.click(timeout=5000)
                    except:
                        print(f"Patient {patient['acct_no']} not found - skipping")
                        continue
                await page.wait_for_timeout(1000)
                await page.locator('[data-test-id="pcl-payments-sendMessageLinkGuarantorDrawer"]').click()
                await page.wait_for_timeout(1000)
                try:
                    await page.get_by_role("button", name=practice).click(timeout=5000)
                    await page.wait_for_timeout(500)
                    await page.get_by_role("menuitem", name="Appointment Scheduling").get_by_role("radio").check(timeout=5000)
                    await page.locator("#menu- > div").first.click(timeout=5000)
                    await page.wait_for_timeout(500)
                except:
                    print("Message type selection skipped")
                message_box = page.get_by_role("textbox", name="Type your response and send")
                await message_box.click()
                await message_box.fill(PCARELINK_MESSAGE)
                print("Message typed!")
                await page.locator('[data-test-id="pcl-payments-sendMessageButton"]').click()
                print("Message sent!")
                await page.wait_for_timeout(1000)
                await page.locator('[data-test-id="pcl-appointments-closePatientsDetails"]').click()
                await page.wait_for_timeout(1000)
            except Exception as e:
                print(f"Error: {e}")
                try:
                    await page.locator('[data-test-id="pcl-appointments-closePatientsDetails"]').click()
                except:
                    pass
                continue

        await browser.close()
        print("\nMessages sent!")

# ─────────────────────────────────────────
# STEP 3 — PEDIFORMS: CHECK & DOWNLOAD
# ─────────────────────────────────────────

async def check_and_download_completed(page, patients):
    """
    Search each patient without status filter, then check if they have
    Completed submissions.

    (2026-07-22 fix) Downloads EVERY completed submission for a patient,
    not just the first - previously `export_buttons.first.click()` always
    grabbed only one PDF regardless of how many completed forms existed,
    which silently dropped M-CHAT/TB completions for multi-form patients.
    Each submission is now matched (best-effort, via its row's text)
    against this patient's expected form names - reconstructed from the
    comma-joined `patient["form_name"]` already persisted at send time, no
    state_db schema change needed - and saved under a filename derived
    from THAT specific form's own sanitized name (e.g.
    "Smith_John_ASQ_18_Months.pdf", "Smith_John_M_Chat.pdf",
    "Smith_John_TB.pdf"). If a row's text can't be matched to any expected
    form, it falls back to a generic "completed_form_N" name rather than
    guessing a specific (possibly wrong) label. Filenames are deterministic,
    so a submission already downloaded in a prior run is detected via
    os.path.exists and skipped, rather than re-downloaded.

    A patient is only marked 'downloaded' (state_db.mark_downloaded, which
    hands them to the upload step) once EVERY expected form has a matching
    file already present in their folder - a patient with e.g. ASQ done
    but M-CHAT/TB still pending stays in 'form_sent' and is re-checked
    again next run, rather than being finalized on a partial capture. See
    this function's LIMITATION note at the bottom of the file for what
    this implies if a parent never completes every expected form.

    Single pass only - returns the list of patients newly downloaded
    this run. Cron re-invoking this script provides the "check again
    later" behavior; there is no internal wait loop here anymore.
    """
    print(f"\nChecking for completed forms...")
    newly_downloaded = []
    # Set week filter only — no status filter to avoid missing patients
    # with multiple submissions where overall status may show differently
    await page.get_by_role("combobox").nth(4).select_option("week")
    await page.wait_for_timeout(1000)

    for patient in patients:
        try:
            search_box = page.get_by_role("textbox", name="Search…")
            await search_box.click()
            await search_box.fill(patient["acct_no"])
            await page.wait_for_timeout(2000)

            view_count = await page.get_by_role("link", name="View").count()
            if view_count == 0:
                print(f"Patient {patient['acct_no']} not found this week")
                await search_box.fill("")
                continue

            # Check if Completed badge is visible in the row
            completed_visible = await page.get_by_text("Completed", exact=True).count()
            if completed_visible == 0:
                print(f"Patient {patient['acct_no']} not completed yet")
                await search_box.fill("")
                continue

            print(f"Patient {patient['acct_no']} has Completed form(s) - opening...")
            await page.get_by_role("link", name="View").first.click()
            await page.wait_for_timeout(1000)

            export_buttons = page.get_by_role("button", name="Export PDF")
            export_count = await export_buttons.count()

            if export_count == 0:
                print(f"No Export PDF button found for {patient['acct_no']} - skipping")
                await page.get_by_role("link", name="← Back to today's patients").click()
                await page.wait_for_timeout(1000)
                continue

            folder_path = ensure_patient_folder(patient)
            # Reconstructed from the comma-joined summary built in
            # read_patients_from_excel() - safe to split on "," since
            # individual form names never contain commas (unlike
            # form_filename, which IS joined with "_" and can't be safely
            # reverse-split since individual filenames also contain "_").
            expected_names = [n.strip() for n in patient.get("form_name", "").split(",") if n.strip()]

            print(f"Found {export_count} completed submission(s) for {patient['acct_no']} - downloading all...")
            for i in range(export_count):
                form_label = None
                try:
                    row_text = await export_buttons.nth(i).locator(
                        "xpath=ancestor::*[self::tr or self::div][1]"
                    ).inner_text()
                    for name in expected_names:
                        if name.lower() in row_text.lower():
                            form_label = name
                            break
                except Exception:
                    pass
                if form_label is None:
                    form_label = f"completed form {i + 1}"

                filename_part = re.sub(r"[^A-Za-z0-9]+", "_", form_label).strip("_")
                file_name = f"{patient['last_name']}_{patient['first_name']}_{filename_part}.pdf"
                save_path = os.path.join(folder_path, file_name)

                if os.path.exists(save_path):
                    print(f"Already downloaded: {file_name} - skipping")
                    continue

                try:
                    print(f"Downloading completed form {i + 1}/{export_count} for {patient['acct_no']} ({form_label})...")
                    async with page.expect_download() as download_info:
                        await export_buttons.nth(i).click()
                    download = await download_info.value
                    await download.save_as(save_path)
                    print(f"Saved: {save_path}")
                except Exception as e:
                    print(f"Error downloading submission {i + 1} for {patient['acct_no']}: {e}")
                    continue

            existing_files = os.listdir(folder_path) if os.path.exists(folder_path) else []
            all_captured = all(
                any(re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_").lower() in f.lower() for f in existing_files)
                for name in expected_names
            ) if expected_names else len(existing_files) > 0

            if all_captured:
                state_db.mark_downloaded(patient["acct_no"], patient["appointment_date"])
                newly_downloaded.append(patient)
                print(f"All expected form(s) captured for {patient['acct_no']} - ready for upload.")
            else:
                print(f"Not all expected forms captured yet for {patient['acct_no']} - will re-check next run.")

            await page.get_by_role("link", name="← Back to today's patients").click()
            await page.wait_for_timeout(1000)

        except Exception as e:
            print(f"Error downloading for {patient['acct_no']}: {e}")
            try:
                await page.get_by_role("link", name="← Back to today's patients").click()
            except:
                pass
            continue

    return newly_downloaded

async def pediforms_check_and_download(patients):
    """
    One single pass over `patients` (pending patients pulled from state_db,
    not just today's Excel). Returns list of newly-downloaded patients.
    """
    print("\n" + "="*50)
    print("STEP 3 — PEDIFORMS: CHECKING FOR COMPLETED FORMS")
    print("="*50)

    if not patients:
        print("No patients pending form completion.")
        return []

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=False, slow_mo=300)
            context = await browser.new_context()
            page = await context.new_page()
            print("Logging into Pediforms...")
            await page.goto("https://admin.pediformpro.com/staff/login", timeout=60000, wait_until="domcontentloaded")
            await asyncio.sleep(3)
            await page.get_by_role("textbox", name="Organization name").fill(PEDIFORMS_ORG)
            await page.get_by_role("textbox", name="Email").fill(PEDIFORMS_EMAIL)
            await page.get_by_role("textbox", name="Password").fill(PEDIFORMS_PASSWORD)
            await page.get_by_role("button", name="Sign in").click()
            await page.wait_for_load_state("networkidle")

            newly_downloaded = await check_and_download_completed(page, patients)
            await browser.close()
            return newly_downloaded

    except Exception as e:
        print(f"\nCheck failed: {e}")
        print("Will retry on next scheduled cron run.")
        return []

# ─────────────────────────────────────────
# STEP 4 — ECW: UPLOAD FORMS
# ─────────────────────────────────────────

async def ecw_upload_forms(patients):
    print("\n" + "="*50)
    print("STEP 4 — ECW: UPLOADING FORMS")
    print("="*50)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=200)
        context = await browser.new_context()
        page = await context.new_page()

        print("Opening eCW login page...")
        await page.goto(
            ECW_LOGIN_URL,
            timeout=120000, wait_until="domcontentloaded"
        )
        await asyncio.sleep(3)
        await page.get_by_role("textbox", name="Enter username to continue").fill(ECW_USERNAME)
        await page.get_by_role("button", name="Next").click()
        print("Username entered!")
        await asyncio.sleep(8)
        await page.click('input[type="password"]')
        await page.keyboard.type(ECW_PASSWORD)
        print("Password entered!")
        await page.keyboard.press("Enter")
        print("Login submitted...")

        await page.wait_for_selector('#jellybean-panelLink33', timeout=120000)
        print("Home page loaded!")
        await _wait_for_loading_overlay_gone(page)

        print("Checking for License Alert...")
        dismissed = False
        for i in range(20):
            try:
                if await page.locator("#providerLicense button.clsMyButton").is_visible():
                    await page.click("#providerLicense button.clsMyButton")
                    print("License Alert dismissed!")
                    dismissed = True
                    break
            except:
                pass
            await asyncio.sleep(1)
        if not dismissed:
            print("No License Alert, continuing...")
        await asyncio.sleep(2)

        # Re-verify right before the next click - see
        # _wait_for_loading_overlay_gone's docstring.
        await _wait_for_loading_overlay_gone(page)

        await page.wait_for_selector("#jellybean-panelLink65", timeout=30000)
        await page.locator("#jellybean-panelLink65").click(force=True)
        await page.get_by_role("textbox", name="Last Name, First Name").wait_for(timeout=30000)
        print("Patient search ready!")

        uploaded_ok = []

        for index, patient in enumerate(patients):
            print(f"\nProcessing {index+1}/{len(patients)}: {patient['last_name']} {patient['first_name']}")
            folder_path = os.path.join(DOC_FOLDER, patient["folder_name"])
            if not os.path.exists(folder_path):
                print(f"Folder not found - skipping")
                continue
            all_files = glob.glob(os.path.join(folder_path, "*"))
            all_files.sort(key=os.path.getmtime)
            if not all_files:
                print(f"No files found - skipping")
                continue
            try:
                search_box = page.get_by_role("textbox", name="Last Name, First Name")
                await search_box.wait_for(timeout=30000)
                await search_box.fill(patient['search_name'])
                await page.wait_for_selector("#patientLName1", timeout=30000)
                try:
                    await page.get_by_role("cell", name=patient['last_name'], exact=False).first.click()
                    await page.get_by_text(patient['last_name'], exact=False).first.click()
                    print("Patient selected!")
                except:
                    print("Could not find patient - skipping")
                    await _go_to_search(page)
                    continue
                try:
                    if await page.locator("text=Please select a patient").is_visible():
                        await page.get_by_role("button", name="OK").click()
                        await _go_to_search(page)
                        continue
                except:
                    pass
                await page.wait_for_selector('button:has-text("Patient Docs")', timeout=30000)
                await page.get_by_role("button", name="Patient Docs").click()
                await page.get_by_role("textbox", name="Quick Search").wait_for(timeout=30000)
                await page.get_by_role("textbox", name="Quick Search").fill("chart")
                await page.wait_for_selector('a:has-text("Chart Documents")', timeout=30000)
                await page.locator("a").filter(has_text="Chart Documents").nth(1).click()
                await asyncio.sleep(2)
                existing_docs = []
                try:
                    doc_links = await page.locator('a[id^="patientdocsTreeLink"]').all()
                    for link in doc_links:
                        doc_object = await link.get_attribute('document-object')
                        if doc_object:
                            doc_data = json.loads(doc_object)
                            label = doc_data.get('label', '').strip().lower()
                            if label:
                                existing_docs.append(label)
                except:
                    pass
                new_files = []
                for f in all_files:
                    filename = os.path.splitext(os.path.basename(f))[0].lower()
                    if filename not in existing_docs:
                        new_files.append(f)
                    else:
                        print(f"Already uploaded: {filename}")
                if not new_files:
                    print("All files already uploaded!")
                    uploaded_ok.append(patient)
                    await _go_to_search(page)
                    continue
                for file_index, file_path in enumerate(new_files):
                    print(f"Uploading {file_index+1}/{len(new_files)}: {os.path.basename(file_path)}")
                    if file_index > 0:
                        await page.get_by_role("textbox", name="Quick Search").wait_for(timeout=30000)
                        await page.get_by_role("textbox", name="Quick Search").fill("chart")
                        await page.wait_for_selector('a:has-text("Chart Documents")', timeout=30000)
                        await page.locator("a").filter(has_text="Chart Documents").nth(1).click()
                        await asyncio.sleep(1)
                    async with page.expect_file_chooser() as fc_info:
                        await page.wait_for_selector("#patientdocsBtn4", timeout=30000)
                        await page.locator("#patientdocsBtn4").click()
                    file_chooser = await fc_info.value
                    await file_chooser.set_files(file_path)
                    await asyncio.sleep(1)
                    try:
                        if await page.locator("text=Please select a category").is_visible():
                            await page.get_by_role("button", name="OK").click()
                    except:
                        pass
                    await page.wait_for_selector('button.commonButton:has-text("OK")', timeout=30000)
                    await page.locator('button.commonButton:has-text("OK")').click()
                    await page.wait_for_selector('#btnOk', timeout=30000)
                    await page.locator('#btnOk').click()
                    print(f"File {file_index+1} saved!")
                    await asyncio.sleep(1)
                print(f"All files uploaded!")
                uploaded_ok.append(patient)
                await _go_to_search(page)
            except Exception as e:
                print(f"Error: {e}")
                try:
                    await _go_to_search(page)
                except:
                    pass
                continue

        await browser.close()
        print("\nAll forms uploaded to eCW!")
        return uploaded_ok

async def _go_to_search(page):
    await page.keyboard.press("Escape")
    await asyncio.sleep(1)
    await page.keyboard.press("Escape")
    await asyncio.sleep(1)
    try:
        await page.wait_for_selector("#patient-hubBtn1", timeout=10000)
        await page.locator("#patient-hubBtn1").click()
        await asyncio.sleep(1)
    except:
        pass
    await page.wait_for_selector("#jellybean-panelLink65", timeout=30000)
    await page.locator("#jellybean-panelLink65").click(force=True)
    await page.get_by_role("textbox", name="Last Name, First Name").wait_for(timeout=30000)

# ─────────────────────────────────────────
# MAIN PIPELINE — PRODUCTION
# ─────────────────────────────────────────

async def main():
    print("="*50)
    print("PRODUCTION PIPELINE — ALL ASQ PATIENTS")
    print("="*50)

    await ecw_export_schedule()

    exported_patients = read_patients_from_excel()
    print(f"\nFound {len(exported_patients)} ASQ patients in this export")

    # --- Split into new vs already-processed (by acct_no + appointment_date) ---
    new_patients = [
        p for p in exported_patients
        if not state_db.is_known(p["acct_no"], p["appointment_date"])
    ]
    already_known = len(exported_patients) - len(new_patients)
    print(f"{len(new_patients)} new patient-visits, {already_known} already processed (skipping resend)")

    if new_patients:
        await pediforms_send_forms(new_patients)
        await pcarelink_send_messages(new_patients)
        for p in new_patients:
            state_db.insert_form_sent(p)
    else:
        print("No new patients to send forms to this run.")

    # --- Check ALL pending patients (from DB, not just today's export) ---
    pending = state_db.get_pending_patients()
    print(f"\n{len(pending)} patient-visits pending form completion (across all runs)")

    if pending:
        newly_downloaded = await pediforms_check_and_download(pending)

        # Upload anything downloaded (this run or a prior run's retry)
        to_upload = state_db.get_patients_needing_upload()
        if to_upload:
            uploaded_ok = await ecw_upload_forms(to_upload)
            for p in uploaded_ok:
                state_db.mark_completed(p["acct_no"], p["appointment_date"])
            print(f"\n{len(uploaded_ok)}/{len(to_upload)} uploaded and marked completed.")

    # --- Housekeeping: drop completed records past the retention window ---
    deleted = state_db.cleanup_old_completed(STATE_RETENTION_DAYS)
    if deleted:
        print(f"\nCleaned up {deleted} completed record(s) older than {STATE_RETENTION_DAYS} days.")

    print("\n" + "="*50)
    print("RUN COMPLETE!")
    print("="*50)

if __name__ == "__main__":
    asyncio.run(main())

# import asyncio
# import os
# import glob
# import json
# import openpyxl
# from dotenv import load_dotenv
# from playwright.async_api import async_playwright

# load_dotenv()

# # --- CREDENTIALS ---
# ECW_USERNAME = os.getenv("ECW_USERNAME")
# ECW_PASSWORD = os.getenv("ECW_PASSWORD")
# PEDIFORMS_ORG = os.getenv("PEDIFORMS_ORG")
# PEDIFORMS_EMAIL = os.getenv("PEDIFORMS_EMAIL")
# PEDIFORMS_PASSWORD = os.getenv("PEDIFORMS_PASSWORD")
# PCARELINK_EMAIL = os.getenv("PCARELINK_EMAIL")
# PCARELINK_PASSWORD = os.getenv("PCARELINK_PASSWORD")
# PCARELINK_PRACTICE = os.getenv("PCARELINK_PRACTICE")
# PCARELINK_MESSAGE = os.getenv("PCARELINK_MESSAGE")
# EXCEL_PATH = os.getenv("EXCEL_PATH")
# DOC_FOLDER = os.getenv("ECW_PATIENTS_DOC_FOLDER")

# # --- SETTINGS ---
# CHECK_INTERVAL_MINUTES = 10  # Check Pediforms every 10 minutes
# MAX_WAIT_HOURS = 24  # Stop checking after 24 hours

# VISIT_TYPE_TO_FORM = {
#     "9 MONTH WC": "ASQ9Mos",
#     "12 MONTH WC": "ASQ 12 Months",
#     "1 YR WC": "ASQ 12 Months",
#     "18 MONTH WC": "ASQ 18 Months",
#     "24 MONTH WC": "ASQ 24 Months",
#     "2 YR WC": "ASQ 24 Months",
#     "30 MONTH WC": "ASQ30",
#     "3 YEAR WC": "ASQ 36 Months",
#     "36 MONTH WC": "ASQ 36 Months",
#     "4 YEAR WC": "ASQ 48 Months",
#     "48 MONTH WC": "ASQ 48 Months",
# }

# VISIT_TYPE_TO_FORM_FILENAME = {
#     "9 MONTH WC": "ASQ9Mos",
#     "12 MONTH WC": "ASQ_12_Months",
#     "1 YR WC": "ASQ_12_Months",
#     "18 MONTH WC": "ASQ_18_Months",
#     "24 MONTH WC": "ASQ_24_Months",
#     "2 YR WC": "ASQ_24_Months",
#     "30 MONTH WC": "ASQ30",
#     "3 YEAR WC": "ASQ_36_Months",
#     "36 MONTH WC": "ASQ_36_Months",
#     "4 YEAR WC": "ASQ_48_Months",
#     "48 MONTH WC": "ASQ_48_Months",
# }

# # ─────────────────────────────────────────
# # SHARED HELPERS
# # ─────────────────────────────────────────

# def read_patients_from_excel():
#     wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
#     ws = wb.active
#     headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
#     col = {name: idx for idx, name in enumerate(headers)}
#     patients = []
#     for row in ws.iter_rows(min_row=2, values_only=True):
#         acct_no = row[col["Patient Acct No"]]
#         last_name = row[col["Patient Last Name"]]
#         first_name = row[col["Patient First Name"]]
#         visit_type_raw = row[col["Visit Type"]]
#         if not acct_no:
#             continue
#         visit_type_desc = str(visit_type_raw).split(":")[-1].strip().upper() if visit_type_raw else ""
#         form_name = VISIT_TYPE_TO_FORM.get(visit_type_desc, None)
#         form_filename = VISIT_TYPE_TO_FORM_FILENAME.get(visit_type_desc, "form")
#         patients.append({
#             "acct_no": str(acct_no).strip(),
#             "last_name": str(last_name).strip() if last_name else "",
#             "first_name": str(first_name).strip() if first_name else "",
#             "folder_name": f"{last_name} {first_name}_doc".strip(),
#             "search_name": f"{last_name},{first_name}".strip(),
#             "visit_type": visit_type_desc,
#             "form_name": form_name,
#             "form_filename": form_filename,
#         })
#     return patients

# def ensure_patient_folder(patient):
#     folder_path = os.path.join(DOC_FOLDER, patient["folder_name"])
#     if not os.path.exists(folder_path):
#         os.makedirs(folder_path)
#         print(f"Created folder: {folder_path}")
#     return folder_path

# # ─────────────────────────────────────────
# # STEP 1 — PEDIFORMS: SEND FORMS
# # ─────────────────────────────────────────

# async def pediforms_send_forms(patients):
#     print("\n" + "="*50)
#     print("STEP 1 — PEDIFORMS: SENDING FORMS")
#     print("="*50)

#     async with async_playwright() as p:
#         browser = await p.chromium.launch(headless=False, slow_mo=300)
#         context = await browser.new_context()
#         page = await context.new_page()

#         print("Logging into Pediforms...")
#         await page.goto("https://admin.pediformpro.com/staff/login")
#         await page.get_by_role("textbox", name="Organization name").fill(PEDIFORMS_ORG)
#         await page.get_by_role("textbox", name="Email").fill(PEDIFORMS_EMAIL)
#         await page.get_by_role("textbox", name="Password").fill(PEDIFORMS_PASSWORD)
#         await page.get_by_role("button", name="Sign in").click()
#         await page.wait_for_load_state("networkidle")
#         print("Logged in!")

#         print("Uploading schedule Excel...")
#         await page.get_by_role("button", name="Choose File").set_input_files(EXCEL_PATH)
#         await page.get_by_role("button", name="Import schedule").click()
#         await page.wait_for_load_state("networkidle")
#         print("Schedule imported!")

#         # Set week filter
#         await page.get_by_role("combobox").nth(4).select_option("week")
#         await page.wait_for_timeout(1000)

#         for patient in patients:
#             try:
#                 if not patient["form_name"]:
#                     print(f"No form mapped for {patient['acct_no']} - skipping")
#                     continue

#                 print(f"\nSending form for {patient['acct_no']} ({patient['last_name']} {patient['first_name']})")

#                 search_box = page.get_by_role("textbox", name="Search…")
#                 await search_box.click()
#                 await search_box.fill(patient["acct_no"])
#                 await page.wait_for_timeout(1000)

#                 try:
#                     await page.get_by_role("link", name="View").click(timeout=10000)
#                 except:
#                     print(f"Patient {patient['acct_no']} not found - skipping")
#                     continue

#                 await page.get_by_role("button", name="+ Send a form").click()
#                 await page.wait_for_timeout(500)

#                 try:
#                     await page.locator("label").filter(has_text=patient["form_name"]).first.click(timeout=10000)
#                 except:
#                     print(f"Form '{patient['form_name']}' not found - skipping")
#                     await page.get_by_role("link", name="← Back to today's patients").click()
#                     await page.wait_for_load_state("networkidle")
#                     continue

#                 await page.get_by_role("button", name="Send form").click()
#                 print(f"Sent '{patient['form_name']}' successfully!")

#                 await page.get_by_role("link", name="← Back to today's patients").click()
#                 await page.wait_for_load_state("networkidle")

#             except Exception as e:
#                 print(f"Error: {e}")
#                 try:
#                     await page.get_by_role("link", name="← Back to today's patients").click()
#                     await page.wait_for_load_state("networkidle")
#                 except:
#                     pass
#                 continue

#         await browser.close()
#         print("\nForms sent to all patients!")

# # ─────────────────────────────────────────
# # STEP 2 — PCARELINK: SEND MESSAGES
# # ─────────────────────────────────────────

# async def pcarelink_send_messages(patients):
#     print("\n" + "="*50)
#     print("STEP 2 — PCARELINK: SENDING MESSAGES")
#     print("="*50)

#     async with async_playwright() as p:
#         browser = await p.chromium.launch(headless=False, slow_mo=300)
#         context = await browser.new_context()
#         page = await context.new_page()

#         print("Logging into pcarelink...")
#         await page.goto("https://app.pcarelink.com/login", timeout=60000, wait_until="domcontentloaded")
#         await asyncio.sleep(3)
#         print("Logged in!")

#         await page.get_by_role("textbox", name="Enter email id").fill(PCARELINK_EMAIL)
#         await page.get_by_role("textbox", name="Enter password").fill(PCARELINK_PASSWORD)
#         await page.locator('[data-test-id="pcl-login-signInButton"]').click()
#         await asyncio.sleep(5)
#         print("Login submitted!")

#         if "profile" in page.url:
#             print("Complete-your-profile popup detected - dismissing via ReachMyDr logo...")
#             await page.get_by_text("ReachMyDr", exact=False).first.click()
#             await asyncio.sleep(3)

#         await page.locator('[data-test-id="pcl-menuDropDownComponent"]').click()
#         await page.locator('[data-test-id="pcl-dashboard-popOver1"]').click()
#         await page.wait_for_load_state("networkidle")

#         await page.get_by_role("button", name="Filter by Practice").click()
#         await page.get_by_text(f"{PCARELINK_PRACTICE}Round Rock, us").click()
#         await page.wait_for_timeout(2000)
#         print(f"Filtered by practice: {PCARELINK_PRACTICE}")

#         for patient in patients:
#             try:
#                 print(f"\nSending message for {patient['acct_no']} ({patient['last_name']} {patient['first_name']})")

#                 search_box = page.get_by_role("searchbox", name="Enter patient first name or")
#                 await search_box.click()
#                 await search_box.fill(patient["acct_no"])
#                 await page.wait_for_timeout(2000)

#                 try:
#                     await page.get_by_text(f"{patient['last_name'].upper()}, {patient['first_name'].upper()}").first.click(timeout=10000)
#                 except:
#                     try:
#                         await page.locator(".patient-result, .search-result").first.click(timeout=5000)
#                     except:
#                         print(f"Patient {patient['acct_no']} not found - skipping")
#                         continue

#                 await page.wait_for_timeout(1000)
#                 await page.locator('[data-test-id="pcl-payments-sendMessageLinkGuarantorDrawer"]').click()
#                 await page.wait_for_timeout(1000)

#                 try:
#                     await page.get_by_role("button", name=PCARELINK_PRACTICE).click(timeout=5000)
#                     await page.wait_for_timeout(500)
#                     await page.get_by_role("menuitem", name="Appointment Scheduling").get_by_role("radio").check(timeout=5000)
#                     await page.locator("#menu- > div").first.click(timeout=5000)
#                     await page.wait_for_timeout(500)
#                 except:
#                     print("Message type selection skipped")

#                 message_box = page.get_by_role("textbox", name="Type your response and send")
#                 await message_box.click()
#                 await message_box.fill(PCARELINK_MESSAGE)
#                 print("Message typed!")

#                 await page.locator('[data-test-id="pcl-payments-sendMessageButton"]').click()
#                 print(f"Message sent successfully!")
#                 await page.wait_for_timeout(1000)

#                 await page.locator('[data-test-id="pcl-appointments-closePatientsDetails"]').click()
#                 await page.wait_for_timeout(1000)

#             except Exception as e:
#                 print(f"Error: {e}")
#                 try:
#                     await page.locator('[data-test-id="pcl-appointments-closePatientsDetails"]').click()
#                 except:
#                     pass
#                 continue

#         await browser.close()
#         print("\nMessages sent to all patients!")

# # ─────────────────────────────────────────
# # STEP 3 — PEDIFORMS: CHECK & DOWNLOAD
# # ─────────────────────────────────────────

# async def pediforms_check_and_download(patients):
#     print("\n" + "="*50)
#     print("STEP 3 — PEDIFORMS: CHECKING FOR COMPLETED FORMS")
#     print("="*50)

#     # Track which patients have been downloaded
#     downloaded = set()
#     total = len([p for p in patients if p["form_name"]])
#     max_checks = (MAX_WAIT_HOURS * 60) // CHECK_INTERVAL_MINUTES

#     for check_num in range(max_checks):
#         print(f"\nCheck {check_num+1}/{max_checks} — {len(downloaded)}/{total} forms downloaded")

#         async with async_playwright() as p:
#             browser = await p.chromium.launch(headless=False, slow_mo=300)
#             context = await browser.new_context()
#             page = await context.new_page()

#             print("Logging into Pediforms...")
#             await page.goto("https://admin.pediformpro.com/staff/login")
#             await page.get_by_role("textbox", name="Organization name").fill(PEDIFORMS_ORG)
#             await page.get_by_role("textbox", name="Email").fill(PEDIFORMS_EMAIL)
#             await page.get_by_role("textbox", name="Password").fill(PEDIFORMS_PASSWORD)
#             await page.get_by_role("button", name="Sign in").click()
#             await page.wait_for_load_state("networkidle")

#             # Set completed filter
#             await page.get_by_role("combobox").nth(1).select_option("completed")
#             await page.wait_for_timeout(1000)
#             await page.get_by_role("combobox").nth(4).select_option("week")
#             await page.wait_for_timeout(1000)

#             for patient in patients:
#                 if patient["acct_no"] in downloaded:
#                     continue
#                 if not patient["form_name"]:
#                     continue

#                 try:
#                     search_box = page.get_by_role("textbox", name="Search…")
#                     await search_box.click()
#                     await search_box.fill(patient["acct_no"])
#                     await page.wait_for_timeout(2000)

#                     view_count = await page.get_by_role("link", name="View").count()
#                     if view_count == 0:
#                         print(f"Patient {patient['acct_no']} form not completed yet")
#                         await search_box.fill("")
#                         continue

#                     await page.get_by_role("link", name="View").first.click()
#                     await page.wait_for_timeout(1000)

#                     folder_path = ensure_patient_folder(patient)
#                     file_name = f"{patient['last_name']}_{patient['first_name']}_{patient['form_filename']}.pdf"
#                     save_path = os.path.join(folder_path, file_name)

#                     print(f"Downloading form for {patient['acct_no']}...")
#                     async with page.expect_download() as download_info:
#                         await page.get_by_role("button", name="Export PDF").click()
#                     download = await download_info.value
#                     await download.save_as(save_path)
#                     print(f"Saved: {save_path}")

#                     downloaded.add(patient["acct_no"])

#                     await page.get_by_role("link", name="← Back to today's patients").click()
#                     await page.wait_for_timeout(1000)

#                 except Exception as e:
#                     print(f"Error downloading for {patient['acct_no']}: {e}")
#                     try:
#                         await page.get_by_role("link", name="← Back to today's patients").click()
#                     except:
#                         pass
#                     continue

#             await browser.close()

#         # Check if all done
#         if len(downloaded) >= total:
#             print(f"\nAll {total} forms downloaded!")
#             return True

#         print(f"\n{len(downloaded)}/{total} forms downloaded. Waiting {CHECK_INTERVAL_MINUTES} minutes before next check...")
#         await asyncio.sleep(CHECK_INTERVAL_MINUTES * 60)

#     print(f"\nMax wait time reached. {len(downloaded)}/{total} forms downloaded.")
#     return len(downloaded) > 0

# # ─────────────────────────────────────────
# # STEP 4 — ECW: UPLOAD FORMS
# # ─────────────────────────────────────────

# async def ecw_upload_forms(patients):
#     print("\n" + "="*50)
#     print("STEP 4 — ECW: UPLOADING FORMS")
#     print("="*50)

#     async with async_playwright() as p:
#         browser = await p.chromium.launch(headless=False, slow_mo=200)
#         context = await browser.new_context()
#         page = await context.new_page()

#         print("Opening eCW login page...")
#         await page.goto("https://txsnmbapp.ecwcloud.com/mobiledoc/jsp/webemr/login/newLogin.jsp")
#         await page.get_by_role("textbox", name="Enter username to continue").fill(ECW_USERNAME)
#         await page.get_by_role("button", name="Next").click()
#         await page.get_by_role("textbox", name="Enter Password to continue").fill(ECW_PASSWORD)
#         await page.get_by_role("button", name="Log In").click()
#         print("Login submitted...")

#         await page.wait_for_selector('#jellybean-panelLink33', timeout=120000)
#         print("Home page loaded!")

#         try:
#             await page.wait_for_selector('#load', state='hidden', timeout=120000)
#         except:
#             pass

#         # Handle License Alert
#         for i in range(20):
#             try:
#                 if await page.locator("#providerLicense button.clsMyButton").is_visible():
#                     await page.click("#providerLicense button.clsMyButton")
#                     print(f"License Alert dismissed!")
#                     break
#             except:
#                 pass
#             await asyncio.sleep(1)

#         # Open patient search
#         await page.wait_for_selector("#jellybean-panelLink65", timeout=30000)
#         await page.locator("#jellybean-panelLink65").click(force=True)
#         await page.get_by_role("textbox", name="Last Name, First Name").wait_for(timeout=30000)
#         print("Patient search ready!")

#         for index, patient in enumerate(patients):
#             print(f"\nProcessing {index+1}/{len(patients)}: {patient['last_name']} {patient['first_name']}")

#             # Get files from patient folder
#             folder_path = os.path.join(DOC_FOLDER, patient["folder_name"])
#             if not os.path.exists(folder_path):
#                 print(f"Folder not found - skipping")
#                 continue

#             all_files = glob.glob(os.path.join(folder_path, "*"))
#             all_files.sort(key=os.path.getmtime)

#             if not all_files:
#                 print(f"No files found - skipping")
#                 continue

#             try:
#                 # Search patient
#                 search_box = page.get_by_role("textbox", name="Last Name, First Name")
#                 await search_box.wait_for(timeout=30000)
#                 await search_box.fill(patient['search_name'])
#                 await page.wait_for_selector("#patientLName1", timeout=30000)

#                 # Select patient
#                 try:
#                     await page.get_by_role("cell", name=patient['last_name'], exact=False).first.click()
#                     await page.get_by_text(patient['last_name'], exact=False).first.click()
#                     print("Patient selected!")
#                 except:
#                     print("Could not find patient - skipping")
#                     await _go_to_search(page)
#                     continue

#                 # Handle popup
#                 try:
#                     if await page.locator("text=Please select a patient").is_visible():
#                         await page.get_by_role("button", name="OK").click()
#                         await _go_to_search(page)
#                         continue
#                 except:
#                     pass

#                 # Open Patient Docs
#                 await page.wait_for_selector('button:has-text("Patient Docs")', timeout=30000)
#                 await page.get_by_role("button", name="Patient Docs").click()

#                 # Search Chart Documents
#                 await page.get_by_role("textbox", name="Quick Search").wait_for(timeout=30000)
#                 await page.get_by_role("textbox", name="Quick Search").fill("chart")
#                 await page.wait_for_selector('a:has-text("Chart Documents")', timeout=30000)
#                 await page.locator("a").filter(has_text="Chart Documents").nth(1).click()
#                 await asyncio.sleep(2)

#                 # Get existing docs
#                 existing_docs = []
#                 try:
#                     doc_links = await page.locator('a[id^="patientdocsTreeLink"]').all()
#                     for link in doc_links:
#                         doc_object = await link.get_attribute('document-object')
#                         if doc_object:
#                             doc_data = json.loads(doc_object)
#                             label = doc_data.get('label', '').strip().lower()
#                             if label:
#                                 existing_docs.append(label)
#                 except:
#                     pass

#                 # Filter new files
#                 new_files = []
#                 for f in all_files:
#                     filename = os.path.splitext(os.path.basename(f))[0].lower()
#                     if filename not in existing_docs:
#                         new_files.append(f)
#                     else:
#                         print(f"Already uploaded: {filename}")

#                 if not new_files:
#                     print("All files already uploaded!")
#                     await _go_to_search(page)
#                     continue

#                 # Upload files
#                 for file_index, file_path in enumerate(new_files):
#                     print(f"Uploading {file_index+1}/{len(new_files)}: {os.path.basename(file_path)}")

#                     if file_index > 0:
#                         await page.get_by_role("textbox", name="Quick Search").wait_for(timeout=30000)
#                         await page.get_by_role("textbox", name="Quick Search").fill("chart")
#                         await page.wait_for_selector('a:has-text("Chart Documents")', timeout=30000)
#                         await page.locator("a").filter(has_text="Chart Documents").nth(1).click()
#                         await asyncio.sleep(1)

#                     async with page.expect_file_chooser() as fc_info:
#                         await page.wait_for_selector("#patientdocsBtn4", timeout=30000)
#                         await page.locator("#patientdocsBtn4").click()
#                     file_chooser = await fc_info.value
#                     await file_chooser.set_files(file_path)
#                     await asyncio.sleep(1)

#                     try:
#                         if await page.locator("text=Please select a category").is_visible():
#                             await page.get_by_role("button", name="OK").click()
#                     except:
#                         pass

#                     await page.wait_for_selector('button.commonButton:has-text("OK")', timeout=30000)
#                     await page.locator('button.commonButton:has-text("OK")').click()
#                     await page.wait_for_selector('#btnOk', timeout=30000)
#                     await page.locator('#btnOk').click()
#                     print(f"File {file_index+1} saved!")
#                     await asyncio.sleep(1)

#                 print(f"All files uploaded for {patient['last_name']} {patient['first_name']}!")
#                 await _go_to_search(page)

#             except Exception as e:
#                 print(f"Error: {e}")
#                 try:
#                     await _go_to_search(page)
#                 except:
#                     pass
#                 continue

#         await browser.close()
#         print("\nAll forms uploaded to eCW!")

# async def _go_to_search(page):
#     await page.keyboard.press("Escape")
#     await asyncio.sleep(1)
#     await page.keyboard.press("Escape")
#     await asyncio.sleep(1)
#     try:
#         await page.wait_for_selector("#patient-hubBtn1", timeout=10000)
#         await page.locator("#patient-hubBtn1").click()
#         await asyncio.sleep(1)
#     except:
#         pass
#     await page.wait_for_selector("#jellybean-panelLink65", timeout=30000)
#     await page.locator("#jellybean-panelLink65").click()
#     await page.get_by_role("textbox", name="Last Name, First Name").wait_for(timeout=30000)

# # ─────────────────────────────────────────
# # MAIN PIPELINE
# # ─────────────────────────────────────────

# async def main():
#     print("="*50)
#     print("COMPLETE PEDIFORMS PIPELINE")
#     print("="*50)

#     patients = read_patients_from_excel()
#     print(f"Found {len(patients)} patients in Excel")

#     # STEP 1 - Send forms via Pediforms
#     await pediforms_send_forms(patients)

#     # STEP 2 - Send messages via pcarelink
#     await pcarelink_send_messages(patients)

#     # STEP 3 - Wait and check for completed forms
#     has_downloads = await pediforms_check_and_download(patients)

#     # STEP 4 - Upload to eCW
#     if has_downloads:
#         await ecw_upload_forms(patients)
#     else:
#         print("No forms downloaded - skipping eCW upload")

#     print("\n" + "="*50)
#     print("PIPELINE COMPLETE!")
#     print("="*50)

# if __name__ == "__main__":
#     asyncio.run(main())