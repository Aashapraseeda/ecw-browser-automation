import asyncio
import os
import openpyxl
from dotenv import load_dotenv
from playwright.async_api import async_playwright

load_dotenv()

EXCEL_PATH = os.getenv("EXCEL_PATH")
ORG_NAME = os.getenv("PEDIFORMS_ORG")
EMAIL = os.getenv("PEDIFORMS_EMAIL")
PASSWORD = os.getenv("PEDIFORMS_PASSWORD")
DOC_FOLDER = os.getenv("ECW_PATIENTS_DOC_FOLDER")

VISIT_TYPE_TO_FORM = {
    "9 MONTH WC": "ASQ9Mos",
    "12 MONTH WC": "ASQ_12_Months",
    "1 YR WC": "ASQ_12_Months",
    "18 MONTH WC": "ASQ_18_Months",
    "24 MONTH WC": "ASQ_24_Months",
    "2 YR WC": "ASQ_24_Months",
    "30 MONTH WC": "ASQ30",
    "3 YEAR WC": "ASQ_36_Months",
    "36 MONTH WC": "ASQ_36_Months",
    "4 YEAR WC": "ASQ_48_Months",
    "48 MONTH WC": "ASQ_48_Months",
}

def read_patients_from_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}
    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acct_no = row[col["Patient Acct No"]]
        last_name = row[col["Patient Last Name"]]
        first_name = row[col["Patient First Name"]]
        visit_type_raw = row[col["Visit Type"]]
        if not acct_no:
            continue

        # Get form name from visit type
        visit_type_desc = str(visit_type_raw).split(":")[-1].strip().upper() if visit_type_raw else ""
        form_name = VISIT_TYPE_TO_FORM.get(visit_type_desc, "form")

        patients.append({
            "acct_no": str(acct_no).strip(),
            "last_name": str(last_name).strip() if last_name else "",
            "first_name": str(first_name).strip() if first_name else "",
            "folder_name": f"{last_name} {first_name}_doc".strip(),
            "form_name": form_name,
        })
    return patients

def ensure_patient_folder(patient):
    """Create patient folder if it doesn't exist"""
    folder_path = os.path.join(DOC_FOLDER, patient["folder_name"])
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Created folder: {folder_path}")
    else:
        print(f"Folder exists: {folder_path}")
    return folder_path

async def download_form_for_patient(page, patient):
    acct_no = patient["acct_no"]
    print(f"\nProcessing patient acct {acct_no} ({patient['last_name']} {patient['first_name']})")

    # --- SEARCH PATIENT ---
    search_box = page.get_by_role("textbox", name="Search…")
    await search_box.click()
    await search_box.fill(acct_no)
    await page.wait_for_timeout(2000)

    # --- CHECK IF PATIENT HAS COMPLETED FORMS ---
    try:
        view_link = page.get_by_role("link", name="View")
        count = await view_link.count()
        if count == 0:
            print(f"No completed forms found for {acct_no} - skipping")
            await search_box.fill("")
            return
    except:
        print(f"Could not check for patient {acct_no} - skipping")
        return

    # --- CLICK VIEW ---
    await view_link.first.click()
    await page.wait_for_timeout(1000)
    print("Opened patient form!")

    # --- CREATE PATIENT FOLDER ---
    folder_path = ensure_patient_folder(patient)

    # --- BUILD FILENAME ---
    file_name = f"{patient['last_name']}_{patient['first_name']}_{patient['form_name']}.pdf"
    save_path = os.path.join(folder_path, file_name)

    # --- DOWNLOAD EXPORT PDF ---
    print(f"Downloading form as: {file_name}")
    async with page.expect_download() as download_info:
        await page.get_by_role("button", name="Export PDF").click()
    download = await download_info.value
    await download.save_as(save_path)
    print(f"Saved to: {save_path}")

    # --- GO BACK TO PATIENT LIST ---
    await page.get_by_role("link", name="← Back to today's patients").click()
    await page.wait_for_timeout(1000)
    print("Back to patient list!")

async def main():
    patients = read_patients_from_excel()
    print(f"Found {len(patients)} patients in Excel")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=300)
        context = await browser.new_context()
        page = await context.new_page()

        # --- LOGIN ---
        print("Logging into Pediforms...")
        await page.goto("https://admin.pediformpro.com/staff/login")
        await page.get_by_role("textbox", name="Organization name").fill(ORG_NAME)
        await page.get_by_role("textbox", name="Email").fill(EMAIL)
        await page.get_by_role("textbox", name="Password").fill(PASSWORD)
        await page.get_by_role("button", name="Sign in").click()
        await page.wait_for_load_state("networkidle")
        print("Logged in!")

        # --- SET STATUS FILTER TO COMPLETED ---
        print("Setting status filter to downloaded...")
        await page.get_by_role("combobox").nth(1).select_option("Downloaded")
        await page.wait_for_timeout(1000)
        print("Filter set to downloaded!")
    
        # --- SET DATE RANGE TO WEEK ---
        print("Setting date range to week...")
        await page.get_by_role("combobox").nth(4).select_option("week")
        await page.wait_for_timeout(1000)
        print("Date range set to week!")

        # --- PROCESS EACH PATIENT ---
        for patient in patients:
            try:
                await download_form_for_patient(page, patient)
            except Exception as e:
                print(f"Error processing {patient['acct_no']}: {e}")
                try:
                    await page.get_by_role("link", name="← Back to today's patients").click()
                    await page.wait_for_timeout(1000)
                except:
                    pass
                continue

        print("\nAll patients processed!")
        print("Now run ecw_upload_docs.py to upload the downloaded forms to eCW!")
        await page.wait_for_timeout(60000)
        await browser.close()

if __name__ == "__main__":
    asyncio.run(main())