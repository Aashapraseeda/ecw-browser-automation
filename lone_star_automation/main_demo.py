"""
main_demo.py
-------------
Lone Star Pediatrics DEMO pipeline - test patients only.

DEMO-ONLY ARCHITECTURE (confirmed does NOT apply to main.py/production
for now): eligibility + form selection is determined from the eCW
Excel's own Visit Type column (e.g. "3 YEAR WC" -> ASQ 36 Months),
exactly like the reference project's main_1.py - NOT from Patient Forms
Now's own "Visit type" column. That turned out to be a PFN-internal
patient-status field (observed showing "New patient" for a real Well
Check test appointment) unrelated to eCW's well-child-visit coding, not
a translation of it. Patient Forms Now is used only to locate each
patient (search by account number) and send the form.

main.py (production) is untouched and still uses the PFN-table/DOB-based
approach in patient_forms_now.form_sender.determine_and_send_from_pfn_table()
- this file no longer calls that function at all.

Also demo-only (this file, plus a small ECW_automation/main_1.py-style
data flow):
  1. eCW export uses a shorter 3-day window (window_days=2) when
     TESTING_SKIP_ECW_EXPORT is off - matching reference main_1.py.
  2. TESTING_SKIP_ECW_EXPORT: skip the live eCW login/export and read a
     fixed test Excel instead (see module-level constants below).
  3. TESTING_RESET_DEMO_STATE: reset known test patients' state_db
     records at the start of each run, so repeated testing isn't
     blocked by "already processed" bookkeeping from a prior run.

PCareLink messaging IS wired in here too (per explicit confirmation),
so a full demo run also validates the reminder-message step - only run
this against test patients with safe/controlled contact numbers on file.
"""

import asyncio
import os

import openpyxl
from playwright.async_api import async_playwright

import ecw.schedule_export as schedule_export
import ecw.chart_upload as chart_upload
import patient_forms_now.form_sender as form_sender
import patient_forms_now.form_downloader as form_downloader
import pcarelink.messenger as pcarelink_messenger
from config import settings
from database import state_db
from patient_forms_now.login import pfn_login
from patient_forms_now.schedule_import import import_schedule
from utils.logger import get_logger

log = get_logger(__name__)

DEMO_WINDOW_DAYS = 2  # today -> today+2 (3-day span), matching reference main_1.py

# TEMPORARY TESTING SWITCH (this file only - main.py is untouched): skip
# the live eCW login/export and read a fixed test Excel instead, so the
# PFN/PCareLink/download/upload steps can be iterated on quickly without
# waiting through the eCW report every time. Set back to False (the
# normal/default behavior) to resume starting from a fresh eCW export.
TESTING_SKIP_ECW_EXPORT = True
TESTING_EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_patient_schedule.xlsx")

# TEMPORARY TESTING SWITCH (this file only - main.py's state tracking is
# untouched): reset settings.DEMO_TEST_ACCOUNT_NUMBERS' state_db records
# at the start of each run, so repeated testing against the same patients
# isn't blocked by "already processed" bookkeeping from a prior run. Set
# to False to restore normal state-tracking for this demo file too.
TESTING_RESET_DEMO_STATE = True

# DEMO-ONLY: eligibility + form selection from the eCW Visit Type text.
# Form label TEXT corrected from a live screenshot of Lone Star's own PFN
# "+ Send a form" checkbox list, which uses a DIFFERENT naming convention
# than the reference clinic's Pediforms account (mostly hyphenated -
# "ASQ-36 Months" - except 48-month, which is "ASQ 48 Months" with a
# space, no hyphen). main.py (production) does not use this dict - it
# still determines eligibility from the PFN table itself via
# form_sender.determine_and_send_from_pfn_table(), which reads
# config.settings.ASQ_BRACKET_TO_FORM (also corrected to match).
VISIT_TYPE_TO_FORM = {
    "9 MONTH WC": "ASQ-9 Months",
    "12 MONTH WC": "ASQ-12 Months",
    "12 MONTHWC": "ASQ-12 Months",
    "1 YEAR WC": "ASQ-12 Months",
    "15 MONTH WC": "ASQ-18 Months",
    "15 MONTHWC": "ASQ-18 Months",
    "18 MONTH WC": "ASQ-18 Months",
    "18 MONTHWC": "ASQ-18 Months",
    "24 MONTH WC": "ASQ-24 Months",
    "24 MONTHWC": "ASQ-24 Months",
    "2 YEAR WC": "ASQ-24 Months",
    "30 MONTH WC": "ASQ-30 Months",
    "30 MONTHWC": "ASQ-30 Months",
    "3 YEAR WC": "ASQ-36 Months",
    "36 MONTH WC": "ASQ-36 Months",
    "36 MONTHWC": "ASQ-36 Months",
    "4 YEAR WC": "ASQ 48 Months",
    "48 MONTH WC": "ASQ 48 Months",
    "48 MONTHWC": "ASQ 48 Months",
}

VISIT_TYPE_TO_FORM_FILENAME = {
    "9 MONTH WC": "ASQ_9_Months",
    "12 MONTH WC": "ASQ_12_Months",
    "12 MONTHWC": "ASQ_12_Months",
    "1 YEAR WC": "ASQ_12_Months",
    "15 MONTH WC": "ASQ_18_Months",
    "15 MONTHWC": "ASQ_18_Months",
    "18 MONTH WC": "ASQ_18_Months",
    "18 MONTHWC": "ASQ_18_Months",
    "24 MONTH WC": "ASQ_24_Months",
    "24 MONTHWC": "ASQ_24_Months",
    "2 YEAR WC": "ASQ_24_Months",
    "30 MONTH WC": "ASQ_30_Months",
    "30 MONTHWC": "ASQ_30_Months",
    "3 YEAR WC": "ASQ_36_Months",
    "36 MONTH WC": "ASQ_36_Months",
    "36 MONTHWC": "ASQ_36_Months",
    "4 YEAR WC": "ASQ_48_Months",
    "48 MONTH WC": "ASQ_48_Months",
    "48 MONTHWC": "ASQ_48_Months",
}


def _build_facility_filtered_excel(source_path):
    """
    TESTING-ONLY HELPER: TESTING_EXCEL_PATH is a shared multi-facility test
    file (also contains River Ridge/Nurture Kids test patients) - importing
    it as-is into Patient Forms Now would upload OTHER clinics' patients
    into Lone Star's own PFN account. In normal/live mode this filtering
    isn't needed, because the eCW export's own Facility filter
    (ecw/facility_filter.py) already guarantees the exported Excel only
    ever contains Lone Star Midlothian rows before this function would run.
    Here, we filter down to Lone Star rows only before import, mirroring
    what main_1.py's FILTERED_EXCEL_PATH already does for the reference
    clinic side.
    """
    wb = openpyxl.load_workbook(source_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}

    if "Appointment Facility Name" not in col:
        log.info("Test Excel has no 'Appointment Facility Name' column - cannot filter, importing as-is")
        return source_path

    own_facility_norm = settings.FACILITY_NAME.strip().lower()
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(headers)
    kept = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        facility_raw = row[col["Appointment Facility Name"]]
        facility_norm = str(facility_raw).strip().lower() if facility_raw else ""
        if facility_norm == own_facility_norm:
            out_ws.append(list(row))
            kept += 1

    out_path = os.path.join(os.path.dirname(source_path), "test_patient_schedule_lonestar_only.xlsx")
    out_wb.save(out_path)
    log.info(f"Filtered test Excel to {kept} Lone Star Midlothian row(s): {out_path}")
    return out_path


def read_demo_patients_from_excel(excel_path):
    """
    DEMO-ONLY: determines eligibility + form purely from the eCW Visit
    Type column (VISIT_TYPE_TO_FORM above) - NOT from PFN, NOT from DOB.
    Restricted to settings.DEMO_TEST_ACCOUNT_NUMBERS as an explicit safety
    gate (same allowlist used elsewhere in this project).
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}
    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acct_no = row[col["Patient Acct No"]]
        last_name = row[col["Patient Last Name"]]
        first_name = row[col["Patient First Name"]]
        visit_type_raw = row[col["Visit Type"]]
        appointment_date_raw = row[col["Appointment Date"]]
        facility_raw = row[col["Appointment Facility Name"]] if "Appointment Facility Name" in col else None
        if not acct_no:
            continue
        acct_no_norm = str(acct_no).strip()
        if acct_no_norm not in settings.DEMO_TEST_ACCOUNT_NUMBERS:
            continue
        visit_type_desc = str(visit_type_raw).split(":")[-1].strip().upper() if visit_type_raw else ""
        form_name = VISIT_TYPE_TO_FORM.get(visit_type_desc, None)
        form_filename = VISIT_TYPE_TO_FORM_FILENAME.get(visit_type_desc, "form")
        if not form_name:
            log.info(f"Skipping {last_name} {first_name} - no ASQ form for visit type: {visit_type_desc!r}")
            continue
        patients.append({
            "acct_no": acct_no_norm,
            "appointment_date": state_db.normalize_date(appointment_date_raw),
            "last_name": str(last_name).strip() if last_name else "",
            "first_name": str(first_name).strip() if first_name else "",
            "folder_name": f"{last_name} {first_name}_doc".strip(),
            "search_name": f"{last_name},{first_name}".strip(),
            "visit_type": visit_type_desc,
            # (2026-07-22) wrapped in a single-element "forms" list purely
            # for compatibility with form_sender's now-shared multi-form
            # send mechanics (_send_forms_for_open_patient) - demo's own
            # Visit-Type-text-based form CHOICE is unchanged, still exactly
            # one ASQ form, no M-CHAT/TB logic here.
            "forms": [{"form_name": form_name, "form_filename": form_filename}],
            "form_name": form_name,
            "form_filename": form_filename,
            "facility": str(facility_raw).strip() if facility_raw else "",
        })
    return patients


async def send_forms_via_excel_visit_type(page, patients):
    """
    DEMO-ONLY: PFN is used only to locate each patient (search by account
    number) and send the form already determined from the Excel - reuses
    form_sender._send_forms_for_open_patient() for the actual send step
    (identical "+ Send a form" -> checkbox -> "Send form" flow used by
    the PFN-table approach, now shared with production's multi-form path).
    """
    sent_patients = []
    for patient in patients:
        try:
            log.info(f"Sending form for {patient['acct_no']} ({patient['last_name']} {patient['first_name']})")
            search_box = page.get_by_role("textbox", name="Search…")
            await search_box.click()
            await search_box.fill(patient["acct_no"])
            await page.wait_for_timeout(1000)
            try:
                await page.get_by_role("link", name="View").click(timeout=10000)
            except Exception:
                log.info(f"Patient {patient['acct_no']} not found in PFN - skipping")
                continue
            sent_forms = await form_sender._send_forms_for_open_patient(page, patient)
            if sent_forms:
                sent_patients.append(patient)
        except Exception as e:
            log.info(f"Error: {e}")
            continue
    return sent_patients


async def run_demo_send(patients):
    """Owns the browser session for the demo Excel-Visit-Type-based send flow."""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=300)
        context = await browser.new_context()
        page = await context.new_page()

        await pfn_login(page)
        await import_schedule(page)
        sent_patients = await send_forms_via_excel_visit_type(page, patients)

        await browser.close()
    return sent_patients


async def main():
    log.info("=" * 50)
    log.info("LONE STAR PEDIATRICS - DEMO PIPELINE (TEST PATIENTS ONLY)")
    log.info("=" * 50)

    if TESTING_RESET_DEMO_STATE:
        deleted = state_db.delete_by_acct_no(settings.DEMO_TEST_ACCOUNT_NUMBERS)
        if deleted:
            log.info(f"TESTING MODE: reset {deleted} prior state record(s) for demo test patients")

    if TESTING_SKIP_ECW_EXPORT:
        log.info("TESTING MODE: skipping eCW login/export - reading fixed test Excel instead:")
        log.info(f"  {TESTING_EXCEL_PATH}")
        # The test Excel is shared across clinics (also has River Ridge/
        # Nurture Kids test patients) - filter to Lone Star's own rows
        # only, so only Lone Star's patient ever gets imported into Lone
        # Star's PFN. Not needed in live mode - the eCW export's own
        # Facility filter already guarantees this upstream.
        settings.EXCEL_PATH = _build_facility_filtered_excel(TESTING_EXCEL_PATH)
    else:
        # --- STEP 0: eCW export (with Facility filter), shorter demo window ---
        await schedule_export.run(window_days=DEMO_WINDOW_DAYS)

    exported_patients = read_demo_patients_from_excel(settings.EXCEL_PATH)
    log.info(f"Found {len(exported_patients)} demo test patients in this export")

    new_patients = [
        p for p in exported_patients
        if not state_db.is_known(p["acct_no"], p["appointment_date"])
    ]
    already_known = len(exported_patients) - len(new_patients)
    log.info(f"{len(new_patients)} new patient-visits, {already_known} already processed (skipping resend)")

    # --- STEP 1: Patient Forms Now - import full schedule, search + send
    # forms already determined from the Excel's Visit Type ---
    sent_patients = []
    if new_patients:
        sent_patients = await run_demo_send(new_patients)
        for p in sent_patients:
            state_db.insert_form_sent(p)

        # --- STEP 2: PCareLink - reminder messages for patients just sent a form ---
        if sent_patients:
            await pcarelink_messenger.send_messages(sent_patients)
    else:
        log.info("No new demo patients to send forms to this run.")

    # --- STEP 3: check ALL pending patients (from DB, not just today's export) ---
    pending = state_db.get_pending_patients()
    log.info(f"{len(pending)} patient-visits pending form completion (across all runs)")

    if pending:
        await form_downloader.run(pending)

        # --- STEP 4: upload anything downloaded (this run or a prior retry) ---
        to_upload = state_db.get_patients_needing_upload()
        if to_upload:
            uploaded_ok = await chart_upload.run(to_upload)
            for p in uploaded_ok:
                state_db.mark_completed(p["acct_no"], p["appointment_date"])
            log.info(f"{len(uploaded_ok)}/{len(to_upload)} uploaded and marked completed.")

    # --- Housekeeping: drop completed records past the retention window ---
    deleted = state_db.cleanup_old_completed(settings.STATE_RETENTION_DAYS)
    if deleted:
        log.info(f"Cleaned up {deleted} completed record(s) older than {settings.STATE_RETENTION_DAYS} days.")

    log.info("=" * 50)
    log.info("DEMO RUN COMPLETE!")
    log.info("=" * 50)


if __name__ == "__main__":
    asyncio.run(main())
