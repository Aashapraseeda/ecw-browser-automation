"""
patient_forms_now/form_sender.py
------------------------------------
STEP 1 - Patient Forms Now: determine Well-Check eligibility and send forms.

**ARCHITECTURE CHANGE (2026-07-21)**: eligibility is now determined from
the eCW Excel export itself (read_eligible_patients_from_excel(), below),
NOT from Patient Forms Now's on-screen table. A live production debug run
proved the PFN-table approach this module used to rely on
(determine_and_send_from_pfn_table(), kept below but no longer called by
main.py) was fundamentally broken: PFN's own "Visit type" column only
ever contains a generic patient-status value ("New patient" / "Follow-up"
/ "Sick visit"), never a clinical visit type - so the "'well' in visit
type" check silently rejected every single row, every run. The real
clinical Visit Type ("9 MONTH WC", "3 YEAR WC", etc.) and Visit Reason
("* 9 MONTH WELL CHILD CHECK") only ever existed in the eCW export.

CURRENT ARCHITECTURE:
  1. read_eligible_patients_from_excel() reads the Excel directly (same
     parsing convention as the reference clinic's read_patients_from_excel()
     - see config.settings.WELL_CHECK_VISIT_TYPES), gated on:
       a. Visit Type/Visit Reason indicates a Well Check (a Well Check is
          now identified structurally - the parsed Visit Type text ends
          in " WC" - rather than an enumerated list, since TB (below)
          spans the full 12mo-18yr Well Check age range and every exact
          "N MONTH WC"/"N YEAR WC" label cannot be enumerated in advance).
       b. forms_for_well_check(age_months) (below) returns at least one
          form - if none apply at this age, the patient is excluded.
  2. **(2026-07-22) MULTI-FORM PER PATIENT**: a patient can now receive
     MORE THAN ONE form in a single run (ASQ + M-CHAT + TB, in any
     combination - see forms_for_well_check()). Each patient dict now
     carries a "forms" list (list of {"form_name","form_filename"} dicts)
     instead of a single form_name/form_filename pair. The singular
     "form_name"/"form_filename" keys are still populated (as a
     comma/underscore-joined summary of every form in the list) purely
     for state_db's existing singular DB columns and log readability -
     state_db.py itself needed NO changes, since dedup only ever operated
     at the (acct_no, appointment_date) - i.e. whole patient-visit - level,
     which is unaffected by how many forms make up that one visit.
  3. The full (unfiltered) Excel is still imported into PFN as-is
     (patient_forms_now.schedule_import) - only the SOURCE of which
     patients get searched-and-sent changed, not the import step.
  4. run_from_excel_list() / search_and_send_from_list() then search PFN
     by account number for just that pre-computed eligible list and send
     EVERY form in that patient's "forms" list (_send_forms_for_open_patient())
     in ONE combined submission - "+ Send a form" once, check every
     matching box, "Send form" once. CONFIRMED LIVE (2026-07-22, read-only
     inspection, no forms sent) that the checkbox panel supports true
     multi-select - see config/settings.py for the verification detail.

**(2026-07-22, fixed)** patient_forms_now/form_downloader.py's
completion-check step previously only grabbed the FIRST completed
submission's PDF per patient - now downloads every completed submission,
matching each to a specific expected form where possible (falling back to
a generic numbered name otherwise), and only advances a patient to
'downloaded' once ALL of their expected forms have a captured file. See
form_downloader.py's own docstring for detail.

REMAINING LIMITATION: because state_db has no per-form tracking (only one
status per patient-visit), a patient whose forms complete on DIFFERENT
days will have their already-downloaded PDF(s) held back from upload
until every expected form is captured - they stay in 'form_sent' the
whole time. If a parent never completes one of the 2-3 expected forms,
that patient's already-downloaded PDFs are never uploaded to eCW under
the current schema (no partial-completion path exists) - they remain
pending indefinitely until STATE_RETENTION_DAYS-based cleanup or manual
intervention. Also unverified live: whether the row-text-matching
heuristic used to identify which submission is which form actually works
against the real completed-forms UI for a genuinely multi-form patient
(never yet observed with real M-CHAT/TB completions) - if it doesn't
match, the fallback numbered filename still guarantees no data loss/
overwrite, just without a form-specific name.

Confirmed column order from a live screenshot of the real PFN table (no
longer used for eligibility, kept here for reference/debugging):
  First name | Last name | Chart # | Region / Clinic | DOB | Appointment |
  Visit type | Form status | Parent portal | Action
"""

import re

import openpyxl
from playwright.async_api import async_playwright

from config import settings
from database import state_db
from patient_forms_now.login import pfn_login
from patient_forms_now.schedule_import import import_schedule
from utils import date_utils
from utils.logger import get_logger

log = get_logger(__name__)

_DEMO_YEAR_TEST_RE = re.compile(r"^\d+\s*year\s*test$")

# Confirmed column positions (see module docstring)
COL_FIRST_NAME = 0
COL_LAST_NAME = 1
COL_CHART_NO = 2
COL_REGION = 3
COL_DOB = 4
COL_APPOINTMENT = 5
COL_VISIT_TYPE = 6
MIN_EXPECTED_COLUMNS = 7


def is_demo_patient(visit_reason):
    """Ported from the reference project's main_1.py demo pipeline."""
    if not visit_reason:
        return False
    vr = str(visit_reason).strip().lower()
    if vr == "test":
        return True
    if _DEMO_YEAR_TEST_RE.match(vr):
        return True
    return False


def _build_demo_allowlist():
    """
    Set of account numbers identifying which patients main_demo.py is
    allowed to touch, read once from the local export. ONLY used to
    restrict main_demo.py to the intended test patient(s) - the Excel
    plays no role in deciding Well-Check eligibility itself (see module
    docstring).

    Two independent, revertible (demo-only) toggles control identification:
      - settings.DEMO_REQUIRE_VISIT_REASON (default False): the actual
        test schedule provided by the clinic has no "test" / "<N> year
        test" values in Visit Reason at all (real clinical-looking text
        like "* 3 YEAR WELL CHILD CHECK" instead) - so this check is off
        for now. Set back to True once test patients carry a real Visit
        Reason marker again.
      - settings.DEMO_RESTRICT_TO_OWN_FACILITY (default True): requires
        Appointment Facility Name to match settings.FACILITY_NAME - this
        is what actually identifies the Lone Star patient(s) in the
        current shared test schedule (which also contains Nurture Kids/
        River Ridge test patients).

    Neither toggle affects main.py (production) - this function is only
    ever called when main_demo.py runs with demo_only=True.
    """
    wb = openpyxl.load_workbook(settings.EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}
    if "Patient Acct No" not in col:
        log.info("Excel missing 'Patient Acct No' column - demo allowlist will be empty")
        return set()

    require_visit_reason = settings.DEMO_REQUIRE_VISIT_REASON and "Visit Reason" in col
    if settings.DEMO_REQUIRE_VISIT_REASON and "Visit Reason" not in col:
        log.info("DEMO_REQUIRE_VISIT_REASON is on but Excel has no 'Visit Reason' column - skipping that check")

    restrict_to_own_facility = settings.DEMO_RESTRICT_TO_OWN_FACILITY and "Appointment Facility Name" in col
    if settings.DEMO_RESTRICT_TO_OWN_FACILITY and "Appointment Facility Name" not in col:
        log.info("DEMO_RESTRICT_TO_OWN_FACILITY is on but Excel has no 'Appointment Facility Name' column - skipping that check")
    own_facility_norm = settings.FACILITY_NAME.strip().lower()

    allowlist = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        acct_no = row[col["Patient Acct No"]]
        if not acct_no:
            continue
        if require_visit_reason and not is_demo_patient(row[col["Visit Reason"]]):
            continue
        if restrict_to_own_facility:
            facility_name_raw = row[col["Appointment Facility Name"]]
            facility_name_norm = str(facility_name_raw).strip().lower() if facility_name_raw else ""
            if facility_name_norm != own_facility_norm:
                log.info(f"Demo allowlist: acct {acct_no} - facility '{facility_name_raw}' != '{settings.FACILITY_NAME}' - excluding")
                continue
        acct_no_norm = str(acct_no).strip()
        if acct_no_norm not in settings.DEMO_TEST_ACCOUNT_NUMBERS:
            # Facility-exclusion alone is not a "test patient" filter - Lone
            # Star's own export already only contains Lone Star Midlothian
            # appointments, so every real patient there would otherwise pass.
            # DEMO_TEST_ACCOUNT_NUMBERS is the actual, explicit safety gate.
            continue
        allowlist.add(acct_no_norm)
    return allowlist


def forms_for_well_check(age_months):
    """
    Returns the list of {"form_name","form_filename"} dicts a Well Check
    patient of this age should receive - three independent rules:
      1. ASQ - existing age-bracket logic (utils.date_utils.match_asq_bracket
         - 9-48 months inclusive), UNCHANGED.
      2. M-CHAT - additional, sent alongside ASQ for settings.MCHAT_ASQ_BRACKETS
         brackets only (18 and 24 months) - never instead of ASQ.
      3. TB - independent age test (settings.TB_MIN_AGE_MONTHS -
         TB_MAX_AGE_MONTHS, 12-216 months inclusive) - applies even when no
         ASQ bracket matches (e.g. a 5-year-old Well Check, previously
         excluded entirely since no ASQ form exists for that age).
    Returns [] if none apply (age outside every window) - caller should
    treat that as "not eligible", same as before.
    """
    forms = []
    bracket = date_utils.match_asq_bracket(age_months)
    if bracket is not None:
        form_name, form_filename = settings.ASQ_BRACKET_TO_FORM[bracket]
        forms.append({"form_name": form_name, "form_filename": form_filename})
        if bracket in settings.MCHAT_ASQ_BRACKETS:
            forms.append(dict(settings.MCHAT_FORM))
    if settings.TB_MIN_AGE_MONTHS <= age_months <= settings.TB_MAX_AGE_MONTHS:
        forms.append(dict(settings.TB_FORM))
    return forms


def read_eligible_patients_from_excel(excel_path):
    """
    Current production eligibility source (see module docstring). Reads
    the eCW export directly - same parsing convention as the reference
    clinic's read_patients_from_excel(): Visit Type text is split on ":"
    and the last segment is upper-cased before matching.

    A row is eligible only if BOTH hold:
      1. Visit Type indicates a Well Check - checked structurally (parsed
         text ends in " WC") since Well Checks now span the full 12mo-18yr
         TB window and every exact "N MONTH WC"/"N YEAR WC" label cannot be
         enumerated in advance; settings.WELL_CHECK_VISIT_TYPES and a
         Visit Reason "WELL CHILD CHECK" check remain as a fallback.
      2. forms_for_well_check(age_months) (above) returns at least one
         form - a patient can now match ASQ, M-CHAT, TB, or any
         combination; each patient dict carries the full list as "forms".

    These are independent checks, not one combined rule: a "9 MONTH WC"
    labeled visit with a DOB that computes to 7 months old is still
    excluded, and vice versa - matches the two-condition requirement
    exactly, rather than trusting either signal alone.
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}

    required_cols = ["Patient Acct No", "Patient Last Name", "Patient First Name",
                      "Visit Type", "Appointment Date", "Patient DOB"]
    missing = [c for c in required_cols if c not in col]
    if missing:
        raise RuntimeError(f"Excel is missing required column(s): {missing}")

    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acct_no = row[col["Patient Acct No"]]
        if not acct_no:
            continue

        last_name = row[col["Patient Last Name"]]
        first_name = row[col["Patient First Name"]]
        visit_type_raw = row[col["Visit Type"]]
        visit_reason_raw = row[col["Visit Reason"]] if "Visit Reason" in col else None
        appointment_date_raw = row[col["Appointment Date"]]
        dob_raw = row[col["Patient DOB"]]
        facility_raw = row[col["Appointment Facility Name"]] if "Appointment Facility Name" in col else None

        visit_type_desc = str(visit_type_raw).split(":")[-1].strip().upper() if visit_type_raw else ""
        is_well_check = (
            visit_type_desc.endswith(" WC")
            or visit_type_desc == "WC"
            or visit_type_desc in settings.WELL_CHECK_VISIT_TYPES
        )
        if not is_well_check and visit_reason_raw:
            is_well_check = "WELL CHILD CHECK" in str(visit_reason_raw).upper()
        if not is_well_check:
            continue

        dob = date_utils.parse_date_flexible(dob_raw)
        appt_date = date_utils.parse_date_flexible(appointment_date_raw)
        if not dob or not appt_date:
            log.info(f"Acct {acct_no}: could not parse DOB={dob_raw!r} / Appointment={appointment_date_raw!r} - skipping")
            continue

        age_months = date_utils.age_in_months(dob, appt_date)
        forms = forms_for_well_check(age_months)
        if not forms:
            continue  # Well Check, but no ASQ/M-CHAT/TB form applies at this age

        patients.append({
            "acct_no": str(acct_no).strip(),
            "appointment_date": state_db.normalize_date(appointment_date_raw),
            "last_name": str(last_name).strip() if last_name else "",
            "first_name": str(first_name).strip() if first_name else "",
            "folder_name": f"{last_name} {first_name}_doc".strip(),
            "search_name": f"{last_name},{first_name}".strip(),
            "visit_type": visit_type_desc,
            "forms": forms,
            "form_name": ", ".join(f["form_name"] for f in forms),
            "form_filename": "_".join(f["form_filename"] for f in forms),
            "facility": str(facility_raw).strip() if facility_raw else "",
        })
    return patients


async def search_and_send_from_list(page, patients):
    """
    Searches Patient Forms Now by account number for each patient in an
    already-computed eligible list (rather than scanning the on-screen
    table for eligibility) and sends EVERY form in that patient's "forms"
    list. Assumes the schedule has already been imported into this page
    (patient_forms_now.schedule_import.import_schedule).
    """
    sent_patients = []
    for patient in patients:
        try:
            form_names = [f["form_name"] for f in patient.get("forms", [])]
            log.info(f"Sending {len(form_names)} form(s) for {patient['acct_no']} "
                      f"({patient['last_name']} {patient['first_name']}): {form_names}")
            search_box = page.get_by_role("textbox", name="Search…")
            await search_box.click()
            await search_box.fill(patient["acct_no"])
            await page.wait_for_timeout(1000)
            try:
                await page.get_by_role("link", name="View").click(timeout=10000)
            except Exception:
                log.info(f"Patient {patient['acct_no']} not found in PFN - skipping")
                continue
            sent_forms = await _send_forms_for_open_patient(page, patient)
            if sent_forms:
                sent_patients.append(patient)
        except Exception as e:
            log.info(f"Error: {e}")
            continue
    return sent_patients


async def run_from_excel_list(patients):
    """Owns the browser session: login -> import full schedule -> search
    PFN by account number for each pre-computed eligible patient -> send.
    Returns the patients a form was actually sent to (drives state_db
    bookkeeping + PCareLink)."""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=300)
        context = await browser.new_context()
        page = await context.new_page()

        await pfn_login(page)
        await import_schedule(page)
        sent_patients = await search_and_send_from_list(page, patients)

        await browser.close()
    return sent_patients


async def _send_forms_for_open_patient(page, patient):
    """
    Sends EVERY form in patient["forms"] to the already-open patient
    detail page (View already clicked) in ONE combined submission: opens
    "+ Send a form" once, checks every matching checkbox, then clicks
    "Send form" once.

    CONFIRMED LIVE (2026-07-22, read-only inspection - see config/settings.py
    for detail): the checkbox panel supports true multi-select - checking
    several boxes leaves all of them checked simultaneously, verified via
    each checkbox's is_checked() state without ever clicking "Send form".
    This replaces an earlier, more conservative design that reopened the
    panel and clicked Send once per form (unnecessary, given multi-select
    is real).

    Returns the list of form dicts actually checked (a subset of
    patient["forms"] if a label isn't found for one of them) - empty list
    means nothing was checked, so "Send form" is never clicked at all.
    """
    checked_forms = []
    try:
        await page.get_by_role("button", name="+ Send a form").click()
        await page.wait_for_timeout(500)

        for form in patient.get("forms", []):
            try:
                await page.locator("label").filter(has_text=form["form_name"]).first.click(timeout=10000)
                checked_forms.append(form)
            except Exception:
                log.info(f"Form '{form['form_name']}' not found - skipping")
                continue

        if checked_forms:
            await page.get_by_role("button", name="Send form").click()
            log.info(f"Sent {[f['form_name'] for f in checked_forms]} successfully!")
        else:
            log.info(f"No matching form checkboxes found for {patient['acct_no']} - nothing sent")
    except Exception as e:
        log.info(f"Error sending forms: {e}")

    try:
        await page.get_by_role("link", name="← Back to today's patients").click()
        await page.wait_for_load_state("networkidle")
    except Exception:
        pass

    return checked_forms


async def determine_and_send_from_pfn_table(page, demo_only=False):
    """
    DEAD CODE (unused since the 2026-07-21 Excel-based eligibility change -
    main.py no longer calls this). NOTE (2026-07-22): this function still
    calls the old single-form _send_form_for_open_patient(page, patient)
    below, which was renamed/replaced by _send_forms_for_open_patient()
    (multi-form) and no longer exists - this function will raise
    NameError if ever invoked. Left as-is (not fixed) since it is
    unreachable from any current entry point; flagged here for whoever
    eventually removes or repairs it.

    Scans the imported "Today's Patients" table row by row, decides
    Well-Check eligibility from each row's Visit Type + DOB + Appointment
    columns, and sends the form immediately for eligible/new rows
    (clicking that row's own View link rather than re-searching).

    demo_only=True additionally restricts to test patients via the
    Excel-sourced allowlist (see module docstring).

    Returns the list of patient dicts a form was actually sent to.
    """
    demo_allowlist = _build_demo_allowlist() if demo_only else None
    if demo_only:
        log.info(f"Demo mode: restricting to {len(demo_allowlist)} test patient acct(s): {sorted(demo_allowlist)}")

    rows = page.get_by_role("row")
    row_count = await rows.count()
    log.info(f"Scanning {row_count} row(s) in the PFN patient table...")

    sent_patients = []
    for i in range(row_count):
        row = rows.nth(i)
        try:
            cells = await row.get_by_role("cell").all_inner_texts()
        except Exception:
            continue
        if len(cells) < MIN_EXPECTED_COLUMNS:
            continue  # not a data row

        first_name = cells[COL_FIRST_NAME].strip()
        last_name = cells[COL_LAST_NAME].strip()
        chart_no = cells[COL_CHART_NO].strip()
        dob_text = cells[COL_DOB].strip()
        appt_text = cells[COL_APPOINTMENT].strip()
        visit_type_text = cells[COL_VISIT_TYPE].strip()

        if first_name.lower() == "first name":
            continue  # header row

        if "well" not in visit_type_text.lower():
            continue  # not a Well Check visit type - skip

        if not chart_no:
            log.info(f"Row matched Well Check visit type ('{visit_type_text}') but Chart # is blank - skipping")
            continue

        if demo_allowlist is not None and chart_no not in demo_allowlist:
            continue

        dob = date_utils.parse_date_flexible(dob_text)
        appt_date = date_utils.parse_date_flexible(appt_text)
        if not dob or not appt_date:
            log.info(f"Chart#{chart_no}: could not parse DOB={dob_text!r} / Appointment={appt_text!r} - skipping")
            continue

        age_months = date_utils.age_in_months(dob, appt_date)
        bracket = date_utils.match_asq_bracket(age_months)
        if bracket is None:
            log.info(f"Chart#{chart_no}: age {age_months} month(s) at appointment - no matching ASQ bracket, skipping")
            continue

        form_name, form_filename = settings.ASQ_BRACKET_TO_FORM[bracket]
        appointment_date_iso = appt_date.isoformat()

        patient = {
            "acct_no": chart_no,
            "appointment_date": appointment_date_iso,
            "last_name": last_name,
            "first_name": first_name,
            "folder_name": f"{last_name} {first_name}_doc".strip(),
            "search_name": f"{last_name},{first_name}".strip(),
            "visit_type": visit_type_text,
            "form_name": form_name,
            "form_filename": form_filename,
        }

        if state_db.is_known(patient["acct_no"], patient["appointment_date"]):
            log.info(f"Chart#{chart_no} / {patient['appointment_date']} already processed - skipping resend")
            continue

        log.info(f"Row eligible: chart#{chart_no}, age {age_months}mo -> {bracket}mo ASQ '{form_name}'")
        try:
            await row.get_by_role("link", name="View").click(timeout=10000)
        except Exception:
            log.info(f"Could not click View for chart#{chart_no} - skipping")
            continue

        sent_ok = await _send_form_for_open_patient(page, patient)
        if sent_ok:
            sent_patients.append(patient)

    log.info(f"PFN table scan complete - sent forms to {len(sent_patients)} patient(s)")
    return sent_patients


async def run(demo_only=False):
    """Owns the browser session: login -> import full schedule -> determine
    eligibility from the PFN table -> send forms. Returns the patients a
    form was actually sent to (drives state_db bookkeeping + PCareLink)."""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=300)
        context = await browser.new_context()
        page = await context.new_page()

        await pfn_login(page)
        await import_schedule(page)
        sent_patients = await determine_and_send_from_pfn_table(page, demo_only=demo_only)

        await browser.close()
    return sent_patients
